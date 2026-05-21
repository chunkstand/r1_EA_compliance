from __future__ import annotations

from collections import Counter
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZIP_DEFLATED, ZipFile
import hashlib

from openpyxl import Workbook

from .preflight import PreflightFetchResult
from .records import WorkbookSource, normalize_url
from .upstream_evaluation_support import (
    REPO_ROOT,
    _read_jsonl,
    _resolve_case_path,
    _write_json,
    _write_jsonl,
)


def _supplemental_sources(definitions: list[dict]) -> list[WorkbookSource]:
    rows = []
    for index, definition in enumerate(definitions, start=5):
        source_record_id = str(definition["source_record_id"])
        url = str(definition["url"])
        rows.append(
            WorkbookSource(
                source_record_id=source_record_id,
                sheet="R1_Forest_Plan_Document_Register",
                excel_row=index,
                source_id=source_record_id,
                title=str(definition["title"]),
                original_url=url,
                effective_url=url,
                normalized_url=normalize_url(url),
                metadata={
                    "source_input": "r1_forest_plan_document_register",
                    "document_role": str(definition.get("document_role") or "forest_plan_support"),
                    "forest_unit_id": str(definition.get("forest_unit_id") or "flathead-nf"),
                    "review_engine_checks": str(
                        definition.get("review_engine_checks") or "forest plan consistency"
                    ),
                },
            )
        )
    return rows


def _write_batch_runs(*, output_dir: Path, batch_runs: list[dict]) -> None:
    for batch_run in batch_runs:
        batch_run_id = str(batch_run["batch_run_id"])
        run_dir = output_dir / "runs" / batch_run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        ledger_path = run_dir / "batch_ledger.json"
        batches_payload = []
        for batch in batch_run.get("batches", []):
            batch_id = str(batch["batch_id"])
            manifest_path = output_dir / "manifests" / f"download_{batch_id}.jsonl"
            manifest_path.parent.mkdir(parents=True, exist_ok=True)
            records = _materialize_records(output_dir=output_dir, records=batch.get("records") or [])
            _write_jsonl(manifest_path, records)
            batches_payload.append(
                {
                    "batch_id": batch_id,
                    "status": str(batch.get("status") or "passed"),
                    "gate_passed": bool(batch.get("gate_passed", True)),
                    "manifest_path": str(manifest_path),
                    "source_record_ids": [
                        str(value) for value in batch.get("source_record_ids", [])
                    ],
                }
            )
        _write_json(ledger_path, {"run_id": batch_run_id, "batches": batches_payload})
        _write_json(
            run_dir / "summary.json",
            {
                "run_id": batch_run_id,
                "all_passed": bool(batch_run.get("all_passed", True)),
                "ledger_path": str(ledger_path),
            },
        )


def _write_download_run(*, output_dir: Path, scenario: dict) -> None:
    run_id = str(scenario.get("run_id") or "upstream-run")
    run_dir = output_dir / "runs" / run_id
    manifest_dir = output_dir / "manifests"
    run_dir.mkdir(parents=True, exist_ok=True)
    manifest_dir.mkdir(parents=True, exist_ok=True)
    records = _materialize_records(output_dir=output_dir, records=scenario.get("records") or [])
    manifest_path = manifest_dir / f"download_{run_id}.jsonl"
    _write_jsonl(manifest_path, records)
    status_counts = Counter(str(record["status"]) for record in records)
    summary = {
        "run_id": run_id,
        "mode": str(scenario.get("mode") or "download"),
        "filtered_rows": len(records),
        "filtered_override_count": sum(
            1 for record in records if record.get("original_url") != record.get("effective_url")
        ),
        "status_counts": dict(status_counts),
        "manifest_path": str(manifest_path),
        "workbook_sha256": "upstream-eval",
    }
    summary.update(scenario.get("summary_overrides") or {})
    _write_json(run_dir / "summary.json", summary)


def _materialize_records(*, output_dir: Path, records: list[dict]) -> list[dict]:
    materialized = []
    for definition in records:
        record = {
            "source_record_id": str(definition["source_record_id"]),
            "sheet": str(definition.get("sheet") or "Ingest_Checklist"),
            "excel_row": int(definition.get("excel_row", 5)),
            "title": str(definition.get("title") or definition["source_record_id"]),
            "original_url": str(definition["original_url"]),
            "effective_url": str(definition.get("effective_url") or definition["original_url"]),
            "normalized_url": str(
                definition.get("normalized_url")
                or normalize_url(
                    str(definition.get("effective_url") or definition["original_url"])
                )
            ),
            "status": str(definition["status"]),
            "http_status": definition.get("http_status"),
            "fetch_timestamp": definition.get("fetch_timestamp"),
            "duplicate_of": None,
            "validation": {
                "mode": str(definition.get("validation_mode") or "download"),
                "passed": bool(definition.get("validation_passed", True)),
            },
            "metadata": dict(definition.get("metadata") or {}),
        }
        duplicate_of = definition.get("duplicate_of")
        if duplicate_of:
            record["duplicate_of"] = str(output_dir / str(duplicate_of))
        artifact = definition.get("artifact")
        artifact_path = definition.get("artifact_path")
        if artifact_path:
            resolved_artifact_path = output_dir / str(artifact_path)
            record["artifact_path"] = str(resolved_artifact_path)
            if artifact:
                resolved_artifact_path.parent.mkdir(parents=True, exist_ok=True)
                body = _artifact_bytes(artifact)
                resolved_artifact_path.write_bytes(body)
                record["artifact_sha256"] = hashlib.sha256(body).hexdigest()
                record["artifact_byte_size"] = len(body)
            else:
                record["artifact_sha256"] = definition.get("artifact_sha256")
                record["artifact_byte_size"] = definition.get("artifact_byte_size")
            if resolved_artifact_path.exists() and (
                record["artifact_sha256"] is None or record["artifact_byte_size"] is None
            ):
                body = resolved_artifact_path.read_bytes()
                record["artifact_sha256"] = hashlib.sha256(body).hexdigest()
                record["artifact_byte_size"] = len(body)
        else:
            record["artifact_path"] = None
            record["artifact_sha256"] = None
            record["artifact_byte_size"] = None
        if definition.get("content_type"):
            record["content_type"] = str(definition["content_type"])
        if definition.get("final_url"):
            record["final_url"] = str(definition["final_url"])
        materialized.append(record)
    return materialized


def _artifact_bytes(spec: dict) -> bytes:
    artifact_type = str(spec["type"])
    if artifact_type == "html":
        return str(spec["body"]).encode("utf-8")
    if artifact_type == "xml":
        return str(spec["body"]).encode("utf-8")
    if artifact_type == "text":
        return str(spec["body"]).encode("utf-8")
    if artifact_type == "pdf":
        if spec.get("pages"):
            return _pdf_bytes(
                [[str(line) for line in page] for page in spec.get("pages", [])]
            )
        return _pdf_bytes([[str(line) for line in spec.get("lines", [])]])
    if artifact_type == "docx":
        return _docx_bytes([str(value) for value in spec.get("paragraphs", [])])
    raise ValueError(f"Unsupported artifact type: {artifact_type}")


def _write_workbook(path: Path, workbook_spec: dict) -> None:
    workbook = Workbook()
    ingest_sheet = workbook.active
    ingest_sheet.title = "Ingest_Checklist"
    forest_plan_sheet = workbook.create_sheet("R1_Forest_Plans")
    exclusions_sheet = workbook.create_sheet("Scope_Exclusions")

    ingest_headers = [
        "ID",
        "Document / Source",
        "Official_Link",
        "Ingest_Status",
        "Scope",
        "Layer",
        "Issuer",
        "Document_Type",
        "Applies_To",
        "Trigger / When to Apply",
        "Review_Engine_Checks",
        "Currentness / Notes",
    ]
    forest_headers = [
        "Unit / Overlay",
        "Current Document or Source",
        "Official_Link",
        "Status in Apr 2026",
        "Applies_When",
        "Review_Engine_Checks",
        "Notes",
    ]
    exclusion_headers = ["Link"]

    _write_sheet_headers(ingest_sheet, ingest_headers)
    _write_sheet_headers(forest_plan_sheet, forest_headers)
    _write_sheet_headers(exclusions_sheet, exclusion_headers)

    for row_index, row in enumerate(workbook_spec.get("ingest_rows", []), start=5):
        ingest_sheet.cell(row_index, 1).value = row["source_record_id"]
        ingest_sheet.cell(row_index, 2).value = row["title"]
        ingest_sheet.cell(row_index, 3).value = row["url"]
        ingest_sheet.cell(row_index, 4).value = row.get("ingest_status", "Ready")
        ingest_sheet.cell(row_index, 5).value = row.get("scope", "Baseline")
        ingest_sheet.cell(row_index, 6).value = row.get("layer", "Federal")
        ingest_sheet.cell(row_index, 7).value = row.get("issuer", "U.S. Congress")
        ingest_sheet.cell(row_index, 8).value = row.get("document_type", "public law")
        ingest_sheet.cell(row_index, 9).value = row.get("applies_to", "Region 1")
        ingest_sheet.cell(row_index, 10).value = row.get("trigger", "Always")
        ingest_sheet.cell(row_index, 11).value = row.get(
            "review_engine_checks", "upstream eval"
        )
        ingest_sheet.cell(row_index, 12).value = row.get("currentness_notes", "fixture")

    for row_index, row in enumerate(workbook_spec.get("forest_plan_rows", []), start=5):
        forest_plan_sheet.cell(row_index, 1).value = row["unit_or_overlay"]
        forest_plan_sheet.cell(row_index, 2).value = row["title"]
        forest_plan_sheet.cell(row_index, 3).value = row["url"]
        forest_plan_sheet.cell(row_index, 4).value = row.get("status", "Current")
        forest_plan_sheet.cell(row_index, 5).value = row.get("applies_when", "Always")
        forest_plan_sheet.cell(row_index, 6).value = row.get(
            "review_engine_checks", "upstream eval"
        )
        forest_plan_sheet.cell(row_index, 7).value = row.get("notes", "fixture")

    for row_index, url in enumerate(workbook_spec.get("excluded_urls", []), start=5):
        exclusions_sheet.cell(row_index, 1).value = url

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _scenario_workbook_path(temp_dir: Path, scenario: dict) -> Path:
    workbook_path = scenario.get("workbook_path")
    if workbook_path is not None:
        return _resolve_case_path(REPO_ROOT, workbook_path)
    workbook_spec = scenario.get("workbook")
    if not isinstance(workbook_spec, dict):
        raise ValueError("Upstream evaluation scenario must provide workbook or workbook_path")
    path = temp_dir / "workbook.xlsx"
    _write_workbook(path, workbook_spec)
    return path


def _write_sheet_headers(sheet, headers: list[str]) -> None:  # noqa: ANN001
    for row_index in range(1, 4):
        sheet.cell(row_index, 1).value = None
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index).value = header


def _pdf_bytes(pages: list[list[str]]) -> bytes:
    if not pages:
        pages = [[]]

    objects: list[bytes] = []
    page_object_ids: list[int] = []
    next_object_id = 1

    catalog_id = next_object_id
    next_object_id += 1
    pages_id = next_object_id
    next_object_id += 1
    font_id = next_object_id
    next_object_id += 1

    objects.append(b"")
    objects.append(b"")
    objects.append(b"")

    for page_lines in pages:
        content_lines = ["BT", "/F1 12 Tf", "72 720 Td"]
        for index, line in enumerate(page_lines):
            escaped = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            if index:
                content_lines.append("0 -16 Td")
            content_lines.append(f"({escaped}) Tj")
        content_lines.append("ET")
        stream = "\n".join(content_lines).encode("utf-8")

        page_id = next_object_id
        content_id = next_object_id + 1
        next_object_id += 2
        page_object_ids.append(page_id)
        objects.append(
            (
                f"{page_id} 0 obj << /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 612 792] "
                f"/Contents {content_id} 0 R /Resources << /Font << /F1 {font_id} 0 R >> >> >> endobj\n"
            ).encode("ascii")
        )
        objects.append(
            f"{content_id} 0 obj << /Length {len(stream)} >> stream\n".encode("ascii")
            + stream
            + b"\nendstream\nendobj\n"
        )

    objects[catalog_id - 1] = (
        f"{catalog_id} 0 obj << /Type /Catalog /Pages {pages_id} 0 R >> endobj\n"
    ).encode("ascii")
    kids = " ".join(f"{page_id} 0 R" for page_id in page_object_ids)
    objects[pages_id - 1] = (
        f"{pages_id} 0 obj << /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >> endobj\n"
    ).encode("ascii")
    objects[font_id - 1] = (
        f"{font_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n"
    ).encode("ascii")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_start = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        (
            f"trailer << /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n"
            f"startxref\n{xref_start}\n%%EOF\n"
        ).encode("ascii")
    )
    return bytes(pdf)


def _docx_bytes(paragraphs: list[str]) -> bytes:
    document_body = "".join(
        f"<w:p><w:r><w:t>{_xml_escape(text)}</w:t></w:r></w:p>" for text in paragraphs
    )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{document_body}</w:body>"
        "</w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        "</Relationships>"
    )
    with TemporaryDirectory() as tmp:
        archive_path = Path(tmp) / "fixture.docx"
        with ZipFile(archive_path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", content_types)
            archive.writestr("_rels/.rels", rels)
            archive.writestr("word/document.xml", document_xml)
        return archive_path.read_bytes()


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _apply_extraction_mutations(
    *,
    extraction_manifest_path: Path,
    chunks_path: Path,
    mutations: list[dict],
) -> None:
    if not mutations:
        return
    manifest_records = _read_jsonl(extraction_manifest_path)
    chunks = _read_jsonl(chunks_path)
    manifest_by_source = {
        str(record.get("source_record_id") or ""): record for record in manifest_records
    }
    chunks_by_source: dict[str, list[dict]] = {}
    for chunk in chunks:
        source_record_id = str(chunk.get("source_record_id") or "")
        chunks_by_source.setdefault(source_record_id, []).append(chunk)

    for mutation in mutations:
        source_record_id = str(
            mutation.get("source_record_id")
            or manifest_records[0].get("source_record_id")
            or ""
        )
        record = manifest_by_source[source_record_id]
        text_path = Path(record["text_path"])
        action = str(mutation["action"])
        if action == "overwrite_extracted_text":
            text_path.write_text(str(mutation["text"]), encoding="utf-8")
        elif action == "append_extracted_text":
            text_path.write_text(
                text_path.read_text(encoding="utf-8") + str(mutation["text"]),
                encoding="utf-8",
            )
        elif action == "overwrite_first_chunk_text":
            source_chunks = chunks_by_source.get(source_record_id) or []
            if not source_chunks:
                raise ValueError(
                    f"No chunks available for mutation source_record_id={source_record_id}"
                )
            source_chunks[0]["text"] = str(mutation["text"])
        elif action == "mark_reused_existing":
            parser_metadata = dict(record.get("parser_metadata") or {})
            parser_metadata["reused_existing"] = True
            record["parser_metadata"] = parser_metadata
        else:
            raise ValueError(f"Unsupported extraction mutation action: {action}")

    _write_jsonl(extraction_manifest_path, manifest_records)
    _write_jsonl(chunks_path, chunks)


def _preflight_result_from_spec(spec: dict) -> PreflightFetchResult:
    return PreflightFetchResult(
        status=str(spec["status"]),
        http_status=int(spec["http_status"]) if spec.get("http_status") is not None else None,
        final_url=str(spec.get("final_url") or ""),
        redirect_chain=[str(value) for value in spec.get("redirect_chain", [])],
        content_type=str(spec.get("content_type") or ""),
        content_length=int(spec["content_length"])
        if spec.get("content_length") is not None
        else None,
        method=str(spec.get("method") or "HEAD"),
        failure=spec.get("failure"),
        validation={
            "mode": "preflight",
            "passed": bool(spec.get("validation_passed", True)),
            "reason": spec.get("validation_reason"),
        },
        attempt_count=int(spec.get("attempt_count", 1)),
    )


def _default_preflight_ok(url: str) -> PreflightFetchResult:
    return PreflightFetchResult(
        status="preflight_ok",
        http_status=200,
        final_url=url,
        redirect_chain=[],
        content_type="text/html",
        content_length=4096,
        method="HEAD",
        failure=None,
        validation={"mode": "preflight", "passed": True, "reason": None},
        attempt_count=1,
    )
