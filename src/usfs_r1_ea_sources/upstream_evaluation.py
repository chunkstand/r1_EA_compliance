from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZIP_DEFLATED, ZipFile
import hashlib
import json

from openpyxl import Workbook

from .catalog import build_review_catalog
from .config import (
    DEFAULT_CONFIG_PATH,
    LEGACY_WORKBOOK_LOADER_CONTRACT,
    load_config,
)
from .dry_run import run_dry_run
from .extract import build_extraction
from .extraction_accuracy import run_extraction_accuracy_audit
from .preflight import PreflightFetchResult, _classify_response, run_preflight
from .records import WorkbookSource, normalize_url
from .validate_run import validate_run


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_UPSTREAM_EVALUATION_MANIFEST_PATH = REPO_ROOT / "config" / "upstream_evaluation_v1.json"
UPSTREAM_EVALUATION_RESULTS_SCHEMA_VERSION = "upstream-evaluation-results-v0"
VERIFIED_EXTRACTION_ADMISSION_CONTRACT_SCHEMA_VERSION = (
    "verified-extraction-admission-contract-v0"
)


@dataclass(frozen=True)
class UpstreamEvaluationResult:
    summary: dict
    output_path: Path
    report_path: Path


def run_upstream_evaluation(
    *,
    manifest_path: Path = DEFAULT_UPSTREAM_EVALUATION_MANIFEST_PATH,
    output_dir: Path = Path("source_library"),
    results_dir: Path | None = None,
) -> UpstreamEvaluationResult:
    manifest_path = _resolve_path(manifest_path)
    output_dir = _resolve_path(output_dir)
    results_dir = _resolve_path(results_dir or output_dir / "evaluations" / "upstream")
    results_dir.mkdir(parents=True, exist_ok=True)

    manifest = _read_json(manifest_path)
    fixture_roots = [
        _resolve_case_path(manifest_path.parent, root) for root in manifest.get("fixture_roots", [])
    ]
    case_definitions = _normalized_case_definitions(
        manifest_dir=manifest_path.parent,
        case_definitions=sorted(
            manifest.get("cases", []),
            key=lambda item: item.get("case_id", ""),
        ),
    )

    contract_checks, output_expectations = _validate_manifest_contract(
        manifest_path=manifest_path,
        manifest=manifest,
        fixture_roots=fixture_roots,
        case_definitions=case_definitions,
    )

    fixture_cache: dict[Path, dict] = {}
    case_results = [
        _run_case(
            case_definition=case_definition,
            fixture_roots=fixture_roots,
            fixture_cache=fixture_cache,
        )
        for case_definition in case_definitions
    ]

    category_summaries = _category_summaries(
        required_lanes=manifest.get("required_lanes", []),
        case_results=case_results,
    )
    lane_summaries = _lane_summaries(
        required_lanes=manifest.get("required_lanes", []),
        category_summaries=category_summaries,
    )

    summary = {
        "schema_version": UPSTREAM_EVALUATION_RESULTS_SCHEMA_VERSION,
        "manifest_path": str(manifest_path),
        "results_dir": str(results_dir),
        "generated_at": _utc_now(),
        "passed": False,
        "required_lane_count": len(manifest.get("required_lanes", [])),
        "required_category_count": sum(
            len(lane.get("required_category_ids", []))
            for lane in manifest.get("required_lanes", [])
        ),
        "case_count": len(case_results),
        "expected_pass_case_count": sum(
            1 for case_result in case_results if case_result["case_type"] == "expected_pass"
        ),
        "controlled_violation_case_count": sum(
            1 for case_result in case_results if case_result["case_type"] == "controlled_violation"
        ),
        "matched_case_count": sum(
            1 for case_result in case_results if case_result["matched_expectation"]
        ),
        "failed_case_ids": [
            case_result["case_id"]
            for case_result in case_results
            if not case_result["matched_expectation"]
        ],
        "contract_checks": contract_checks,
        "lane_summaries": lane_summaries,
        "category_summaries": category_summaries,
        "cases": case_results,
    }

    output_field_check = {
        "name": "output_schema_fields_present",
        "passed": not [
            field for field in output_expectations if field not in summary and field != "passed"
        ],
        "details": {
            "required_fields": output_expectations,
            "missing_fields": [
                field for field in output_expectations if field not in summary and field != "passed"
            ],
        },
    }
    summary["contract_checks"].append(output_field_check)
    summary["passed"] = all(check["passed"] for check in summary["contract_checks"]) and all(
        case_result["matched_expectation"] for case_result in case_results
    )

    output_path = results_dir / "upstream_evaluation_results.json"
    report_path = results_dir / "upstream_evaluation_report.md"
    summary["output_path"] = str(output_path)
    summary["report_path"] = str(report_path)
    _write_json(output_path, summary)
    report_path.write_text(_markdown_report(summary), encoding="utf-8")
    return UpstreamEvaluationResult(summary=summary, output_path=output_path, report_path=report_path)


def _validate_manifest_contract(
    *,
    manifest_path: Path,
    manifest: dict,
    fixture_roots: list[Path],
    case_definitions: list[dict],
) -> tuple[list[dict], list[str]]:
    checks: list[dict] = []
    checks.append(
        {
            "name": "manifest_schema_version",
            "passed": manifest.get("schema_version") == "upstream-evaluation-v1",
            "details": {
                "path": str(manifest_path),
                "actual": manifest.get("schema_version"),
                "expected": "upstream-evaluation-v1",
            },
        }
    )

    required_lanes = manifest.get("required_lanes", [])
    lane_ids = [str(lane.get("lane_id") or "") for lane in required_lanes]
    checks.append(
        {
            "name": "required_lanes_are_unique",
            "passed": all(lane_ids.count(lane_id) == 1 for lane_id in lane_ids if lane_id),
            "details": {"lane_ids": lane_ids},
        }
    )

    required_categories = {
        str(lane.get("lane_id")): {
            str(category_id) for category_id in lane.get("required_category_ids", [])
        }
        for lane in required_lanes
    }
    minimum_expected_cases = {
        str(lane.get("lane_id")): int(lane.get("minimum_expected_pass_cases_per_category", 1))
        for lane in required_lanes
    }
    minimum_controlled_cases = {
        str(lane.get("lane_id")): int(
            lane.get("minimum_controlled_violation_cases_per_category", 1)
        )
        for lane in required_lanes
    }

    case_keys = [(case.get("lane_id"), case.get("category_id"), case.get("case_id")) for case in case_definitions]
    checks.append(
        {
            "name": "case_ids_are_unique",
            "passed": len({case_id for _, _, case_id in case_keys}) == len(case_definitions),
            "details": {"case_ids": [case.get("case_id") for case in case_definitions]},
        }
    )

    missing_categories = []
    thin_categories = []
    invalid_case_shapes = []
    out_of_tree_fixtures = []
    missing_fixtures = []
    case_count_by_category: dict[tuple[str, str, str], int] = Counter()

    for case in case_definitions:
        lane_id = str(case.get("lane_id") or "")
        category_id = str(case.get("category_id") or "")
        case_type = str(case.get("case_type") or "")
        fixture_path = _resolve_path(case.get("fixture_path"))
        case_id = str(case.get("case_id") or "")
        scenario_id = str(case.get("scenario_id") or "")
        expected = case.get("expected_producer_passed")

        if (
            not lane_id
            or not category_id
            or case_type not in {"expected_pass", "controlled_violation"}
            or not case_id
            or not scenario_id
            or not isinstance(expected, bool)
        ):
            invalid_case_shapes.append(case_id or "<missing-case-id>")

        if lane_id not in required_categories or category_id not in required_categories.get(lane_id, set()):
            missing_categories.append({"lane_id": lane_id, "category_id": category_id, "case_id": case_id})

        if case_type and lane_id and category_id:
            case_count_by_category[(lane_id, category_id, case_type)] += 1

        if not fixture_path.exists():
            missing_fixtures.append(str(fixture_path))
        if fixture_roots and not any(_is_within(fixture_path, root) for root in fixture_roots):
            out_of_tree_fixtures.append(str(fixture_path))

    for lane_id, category_ids in required_categories.items():
        for category_id in sorted(category_ids):
            expected_count = case_count_by_category[(lane_id, category_id, "expected_pass")]
            controlled_count = case_count_by_category[(lane_id, category_id, "controlled_violation")]
            if expected_count < minimum_expected_cases[lane_id] or controlled_count < minimum_controlled_cases[lane_id]:
                thin_categories.append(
                    {
                        "lane_id": lane_id,
                        "category_id": category_id,
                        "expected_pass_case_count": expected_count,
                        "controlled_violation_case_count": controlled_count,
                        "minimum_expected_pass_cases_per_category": minimum_expected_cases[lane_id],
                        "minimum_controlled_violation_cases_per_category": minimum_controlled_cases[lane_id],
                    }
                )

    checks.extend(
        [
            {
                "name": "cases_reference_required_categories",
                "passed": not missing_categories,
                "details": {"invalid_case_categories": missing_categories},
            },
            {
                "name": "required_categories_have_expected_and_controlled_cases",
                "passed": not thin_categories,
                "details": {"thin_categories": thin_categories},
            },
            {
                "name": "fixture_paths_are_in_tracked_roots",
                "passed": not out_of_tree_fixtures,
                "details": {
                    "fixture_roots": [str(root) for root in fixture_roots],
                    "out_of_tree_fixture_paths": out_of_tree_fixtures,
                },
            },
            {
                "name": "fixture_paths_exist",
                "passed": not missing_fixtures,
                "details": {"missing_fixture_paths": missing_fixtures},
            },
            {
                "name": "case_definitions_are_well_formed",
                "passed": not invalid_case_shapes,
                "details": {"invalid_case_ids": invalid_case_shapes},
            },
        ]
    )

    output_expectations = [
        str(field)
        for field in (manifest.get("output_schema") or {}).get("required_summary_fields", [])
    ]
    return checks, output_expectations


def _run_case(
    *,
    case_definition: dict,
    fixture_roots: list[Path],
    fixture_cache: dict[Path, dict],
) -> dict:
    case_id = str(case_definition.get("case_id") or "")
    fixture_path = _resolve_path(case_definition.get("fixture_path"))
    scenario_id = str(case_definition.get("scenario_id") or "")
    expected_producer_passed = bool(case_definition.get("expected_producer_passed"))

    try:
        if fixture_roots and not any(_is_within(fixture_path, root) for root in fixture_roots):
            raise ValueError(f"Fixture path is outside tracked fixture roots: {fixture_path}")
        fixture_payload = fixture_cache.setdefault(fixture_path, _read_json(fixture_path))
        runner_name = str(fixture_payload.get("runner") or "")
        scenarios = fixture_payload.get("scenarios") or {}
        scenario = scenarios.get(scenario_id)
        if not runner_name or not isinstance(scenario, dict):
            raise ValueError(f"Fixture {fixture_path} is missing runner or scenario {scenario_id}")
        runner = _RUNNERS.get(runner_name)
        if runner is None:
            raise ValueError(f"Unsupported upstream evaluation runner: {runner_name}")
        actual_producer_passed, assertion_passed, details = runner(scenario)
        error = None
    except Exception as exc:  # noqa: BLE001
        actual_producer_passed = False
        assertion_passed = False
        details = {}
        error = {"error_class": type(exc).__name__, "error_message": str(exc)}

    matched_expectation = (
        actual_producer_passed == expected_producer_passed and assertion_passed and error is None
    )
    return {
        "case_id": case_id,
        "lane_id": str(case_definition.get("lane_id") or ""),
        "category_id": str(case_definition.get("category_id") or ""),
        "case_type": str(case_definition.get("case_type") or ""),
        "fixture_path": str(fixture_path),
        "scenario_id": scenario_id,
        "expected_producer_passed": expected_producer_passed,
        "actual_producer_passed": actual_producer_passed,
        "assertion_passed": assertion_passed,
        "matched_expectation": matched_expectation,
        "details": details,
        "error": error,
    }


def _category_summaries(*, required_lanes: list[dict], case_results: list[dict]) -> list[dict]:
    cases_by_category: dict[tuple[str, str], list[dict]] = {}
    for case_result in case_results:
        key = (case_result["lane_id"], case_result["category_id"])
        cases_by_category.setdefault(key, []).append(case_result)

    summaries = []
    for lane in required_lanes:
        lane_id = str(lane.get("lane_id") or "")
        for category_id in sorted(str(value) for value in lane.get("required_category_ids", [])):
            cases = cases_by_category.get((lane_id, category_id), [])
            failing_case_ids = [
                case_result["case_id"] for case_result in cases if not case_result["matched_expectation"]
            ]
            summaries.append(
                {
                    "lane_id": lane_id,
                    "category_id": category_id,
                    "case_count": len(cases),
                    "expected_pass_case_count": sum(
                        1 for case_result in cases if case_result["case_type"] == "expected_pass"
                    ),
                    "controlled_violation_case_count": sum(
                        1
                        for case_result in cases
                        if case_result["case_type"] == "controlled_violation"
                    ),
                    "matched_case_count": sum(
                        1 for case_result in cases if case_result["matched_expectation"]
                    ),
                    "passed": bool(cases) and not failing_case_ids,
                    "failing_case_ids": failing_case_ids,
                    "status": "direct_eval_present"
                    if cases and not failing_case_ids
                    else "direct_eval_missing",
                }
            )
    return summaries


def _lane_summaries(*, required_lanes: list[dict], category_summaries: list[dict]) -> list[dict]:
    summaries = []
    for lane in required_lanes:
        lane_id = str(lane.get("lane_id") or "")
        lane_categories = [
            summary for summary in category_summaries if summary["lane_id"] == lane_id
        ]
        failing_case_ids = sorted(
            {
                case_id
                for summary in lane_categories
                for case_id in summary.get("failing_case_ids", [])
            }
        )
        summaries.append(
            {
                "lane_id": lane_id,
                "register_rows": lane.get("register_rows", []),
                "category_count": len(lane_categories),
                "direct_eval_present_category_count": sum(
                    1 for summary in lane_categories if summary["status"] == "direct_eval_present"
                ),
                "passed": lane_categories and not failing_case_ids,
                "status": "direct_eval_present"
                if lane_categories and not failing_case_ids
                else "direct_eval_missing",
                "failing_case_ids": failing_case_ids,
            }
        )
    return summaries


def _run_capture_response_classification(scenario: dict) -> tuple[bool, bool, dict]:
    config = _upstream_eval_config()
    response = scenario["response"]
    classification = _classify_response(
        http_status=int(response["http_status"]),
        final_url=str(response["final_url"]),
        redirect_chain=[str(value) for value in response.get("redirect_chain", [])],
        content_type=str(response["content_type"]),
        content_length=int(response["content_length"]),
        method=str(response["method"]),
        attempt_count=int(response.get("attempt_count", 1)),
        body_sample=str(response["body_sample"]).encode("utf-8"),
        validation=config.validation,
    )
    recorded_result = scenario["recorded_result"]
    actual_producer_passed = (
        classification.status == str(recorded_result["status"])
        and bool(classification.validation.get("passed"))
        == bool(recorded_result["validation_passed"])
    )
    return actual_producer_passed, True, {
        "classification_status": classification.status,
        "classification_validation_passed": bool(classification.validation.get("passed")),
        "recorded_status": recorded_result["status"],
        "recorded_validation_passed": bool(recorded_result["validation_passed"]),
    }


def _run_capture_preflight(scenario: dict) -> tuple[bool, bool, dict]:
    config = _upstream_eval_config(
        loader_contract=str(scenario.get("loader_contract") or LEGACY_WORKBOOK_LOADER_CONTRACT)
    )
    with TemporaryDirectory() as tmp:
        output_dir = Path(tmp) / "source_library"
        workbook_path = _scenario_workbook_path(Path(tmp), scenario)
        fetch_results = {
            url: _preflight_result_from_spec(spec)
            for url, spec in (scenario.get("fetch_results") or {}).items()
        }

        def fetcher(url, network, validation):  # noqa: ANN001
            return fetch_results.get(url, _default_preflight_ok(url))

        result = run_preflight(
            workbook_path=workbook_path,
            output_dir=output_dir,
            config=config,
            run_id=str(scenario.get("run_id") or "upstream-preflight"),
            fetcher=fetcher,
            sleep_fn=lambda _: None,
        )
        records = _read_jsonl(result.manifest_path)
        assertions = scenario["assertions"]
        actual_producer_passed = (
            result.summary.get("checked_url_count") == assertions["checked_url_count"]
            and result.summary.get("duplicate_url_count") == assertions["duplicate_url_count"]
            and result.summary.get("filtered_rows") == assertions["filtered_rows"]
            and sum(1 for record in records if record.get("status") == "duplicate_url")
            == assertions["duplicate_record_count"]
            and all(
                record.get("duplicate_of")
                for record in records
                if record.get("status") == "duplicate_url"
            )
        )
        return actual_producer_passed, True, {
            "checked_url_count": result.summary.get("checked_url_count"),
            "duplicate_url_count": result.summary.get("duplicate_url_count"),
            "filtered_rows": result.summary.get("filtered_rows"),
            "duplicate_record_count": sum(
                1 for record in records if record.get("status") == "duplicate_url"
            ),
        }


def _run_capture_dry_run(scenario: dict) -> tuple[bool, bool, dict]:
    config = _upstream_eval_config(
        loader_contract=str(scenario.get("loader_contract") or LEGACY_WORKBOOK_LOADER_CONTRACT)
    )
    with TemporaryDirectory() as tmp:
        output_dir = Path(tmp) / "source_library"
        workbook_path = _scenario_workbook_path(Path(tmp), scenario)
        supplemental_sources = _supplemental_sources(
            scenario.get("supplemental_sources") or []
        )
        expected_error = str(scenario.get("expected_error_contains") or "")
        try:
            result = run_dry_run(
                workbook_path=workbook_path,
                output_dir=output_dir,
                config=config,
                run_id=str(scenario.get("run_id") or "upstream-dry-run"),
                supplemental_sources=supplemental_sources or None,
                source_delta_input=scenario.get("source_delta_input"),
            )
        except Exception as exc:  # noqa: BLE001
            details = {"error_class": type(exc).__name__, "error_message": str(exc)}
            if expected_error:
                return expected_error in str(exc), True, details
            raise

        records = _read_jsonl(result.manifest_path)
        if expected_error:
            return False, True, {
                "error_class": None,
                "error_message": None,
                "filtered_rows": result.summary.get("filtered_rows"),
                "sheet_names": sorted({record.get("sheet") for record in records}),
            }

        assertions = scenario.get("assertions") or {}
        actual_producer_passed = True
        if "filtered_rows" in assertions:
            actual_producer_passed = actual_producer_passed and (
                result.summary.get("filtered_rows") == assertions["filtered_rows"]
            )
        if "unique_canonical_urls" in assertions:
            actual_producer_passed = actual_producer_passed and (
                result.summary.get("unique_canonical_urls")
                == assertions["unique_canonical_urls"]
            )
        if "required_sheet_names" in assertions:
            actual_producer_passed = actual_producer_passed and (
                sorted({record.get("sheet") for record in records})
                == sorted(assertions["required_sheet_names"])
            )
        return actual_producer_passed, True, {
            "filtered_rows": result.summary.get("filtered_rows"),
            "unique_canonical_urls": result.summary.get("unique_canonical_urls"),
            "sheet_names": sorted({record.get("sheet") for record in records}),
        }


def _run_capture_validate_run(scenario: dict) -> tuple[bool, bool, dict]:
    with TemporaryDirectory() as tmp:
        output_dir = Path(tmp) / "source_library"
        _write_download_run(output_dir=output_dir, scenario=scenario)
        result = validate_run(
            output_dir=output_dir,
            run_id=str(scenario.get("run_id") or "upstream-validate-run"),
        )
        required_check_outcomes = scenario.get("required_check_outcomes") or {}
        assertion_passed = all(
            _check_outcome(result.report, check_name) == bool(expected)
            for check_name, expected in required_check_outcomes.items()
        )
        return result.passed, assertion_passed, {
            "validation_passed": result.passed,
            "required_check_outcomes": {
                check_name: _check_outcome(result.report, check_name)
                for check_name in required_check_outcomes
            },
            "failed_checks": _failed_check_names(result.report),
        }


def _run_catalog_build(scenario: dict) -> tuple[bool, bool, dict]:
    config = _upstream_eval_config(
        loader_contract=str(scenario.get("loader_contract") or LEGACY_WORKBOOK_LOADER_CONTRACT)
    )
    with TemporaryDirectory() as tmp:
        output_dir = Path(tmp) / "source_library"
        workbook_path = _scenario_workbook_path(Path(tmp), scenario)
        supplemental_sources = _supplemental_sources(
            scenario.get("supplemental_sources") or []
        )
        _write_batch_runs(output_dir=output_dir, batch_runs=scenario.get("batch_runs") or [])
        build_kwargs = dict(
            workbook_path=workbook_path,
            output_dir=output_dir,
            config=config,
            config_path=REPO_ROOT / DEFAULT_CONFIG_PATH,
            supplemental_sources=supplemental_sources or None,
            source_delta_input=scenario.get("source_delta_input"),
        )
        if scenario.get("batch_run_id"):
            build_kwargs["batch_run_id"] = str(scenario["batch_run_id"])
        if scenario.get("batch_run_ids"):
            build_kwargs["batch_run_ids"] = [str(value) for value in scenario["batch_run_ids"]]
        result = build_review_catalog(**build_kwargs)
        validation = _read_json(result.validation_path)
        required_check_outcomes = scenario.get("required_check_outcomes") or {}
        assertion_passed = all(
            _check_outcome(validation, check_name) == bool(expected)
            for check_name, expected in required_check_outcomes.items()
        )
        return bool(result.summary.get("validation_passed")), assertion_passed, {
            "validation_passed": bool(result.summary.get("validation_passed")),
            "required_check_outcomes": {
                check_name: _check_outcome(validation, check_name)
                for check_name in required_check_outcomes
            },
            "failed_checks": _failed_check_names(validation),
        }


def _run_extraction_build(scenario: dict) -> tuple[bool, bool, dict]:
    config = _upstream_eval_config(
        loader_contract=str(scenario.get("loader_contract") or LEGACY_WORKBOOK_LOADER_CONTRACT)
    )
    with TemporaryDirectory() as tmp:
        output_dir = Path(tmp) / "source_library"
        workbook_path = _scenario_workbook_path(Path(tmp), scenario)
        _write_download_run(output_dir=output_dir, scenario=scenario["download_run"])
        build_review_catalog(
            workbook_path=workbook_path,
            output_dir=output_dir,
            config=config,
            config_path=REPO_ROOT / DEFAULT_CONFIG_PATH,
            run_id=str(scenario["download_run"].get("run_id") or "upstream-extraction"),
            source_record_ids={str(scenario["source_record_id"])},
        )
        extraction = build_extraction(
            output_dir=output_dir,
            id_filter=str(scenario["source_record_id"]),
        )
        _apply_extraction_mutations(
            extraction_manifest_path=extraction.extraction_manifest_path,
            chunks_path=extraction.chunks_path,
            mutations=scenario.get("mutations") or [],
        )
        contract_path = output_dir / "verified_extraction_contract.json"
        _write_json(
            contract_path,
            {
                "schema_version": VERIFIED_EXTRACTION_ADMISSION_CONTRACT_SCHEMA_VERSION,
                "contracts": [
                    {
                        "contract_id": str(scenario["source_record_id"]),
                        "required_source_record_ids": [str(scenario["source_record_id"])],
                        "require_direct_extraction": True,
                    }
                ],
            },
        )
        audit = run_extraction_accuracy_audit(
            output_dir=output_dir,
            contract_path=contract_path,
        )
        manifest_record = next(
            record
            for record in _read_jsonl(extraction.extraction_manifest_path)
            if record.get("source_record_id") == scenario["source_record_id"]
        )
        text = Path(manifest_record["text_path"]).read_text(encoding="utf-8")
        missing_markers = [
            marker for marker in scenario.get("required_markers", []) if marker not in text
        ]
        forbidden_markers = [
            marker for marker in scenario.get("forbidden_markers", []) if marker in text
        ]
        marker_passed = not missing_markers and not forbidden_markers
        required_check_outcomes = scenario.get("required_check_outcomes") or {}
        assertion_passed = all(
            _check_outcome(audit.summary, check_name) == bool(expected)
            for check_name, expected in required_check_outcomes.items()
        )
        actual_producer_passed = bool(audit.summary.get("passed")) and marker_passed
        return actual_producer_passed, assertion_passed, {
            "audit_passed": bool(audit.summary.get("passed")),
            "required_check_outcomes": {
                check_name: _check_outcome(audit.summary, check_name)
                for check_name in required_check_outcomes
            },
            "missing_markers": missing_markers,
            "forbidden_markers_present": forbidden_markers,
            "failed_checks": _failed_check_names(audit.summary),
        }


_RUNNERS = {
    "capture_dry_run": _run_capture_dry_run,
    "capture_preflight": _run_capture_preflight,
    "capture_response_classification": _run_capture_response_classification,
    "capture_validate_run": _run_capture_validate_run,
    "catalog_build": _run_catalog_build,
    "extraction_build": _run_extraction_build,
}


def _supplemental_sources(definitions: list[dict]) -> list[WorkbookSource]:
    rows = []
    for index, definition in enumerate(definitions, start=5):
        source_record_id = str(definition["source_record_id"])
        url = str(definition["url"])
        rows.append(
            WorkbookSource(
                source_record_id=source_record_id,
                sheet="R1_Forest_Plan_Document_Register",
                excel_row=index,
                source_id=source_record_id,
                title=str(definition["title"]),
                original_url=url,
                effective_url=url,
                normalized_url=normalize_url(url),
                metadata={
                    "source_input": "r1_forest_plan_document_register",
                    "document_role": str(definition.get("document_role") or "forest_plan_support"),
                    "forest_unit_id": str(definition.get("forest_unit_id") or "flathead-nf"),
                    "review_engine_checks": str(
                        definition.get("review_engine_checks") or "forest plan consistency"
                    ),
                },
            )
        )
    return rows


def _write_batch_runs(*, output_dir: Path, batch_runs: list[dict]) -> None:
    for batch_run in batch_runs:
        batch_run_id = str(batch_run["batch_run_id"])
        run_dir = output_dir / "runs" / batch_run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        ledger_path = run_dir / "batch_ledger.json"
        batches_payload = []
        for batch in batch_run.get("batches", []):
            batch_id = str(batch["batch_id"])
            manifest_path = output_dir / "manifests" / f"download_{batch_id}.jsonl"
            manifest_path.parent.mkdir(parents=True, exist_ok=True)
            records = _materialize_records(output_dir=output_dir, records=batch.get("records") or [])
            _write_jsonl(manifest_path, records)
            batches_payload.append(
                {
                    "batch_id": batch_id,
                    "status": str(batch.get("status") or "passed"),
                    "gate_passed": bool(batch.get("gate_passed", True)),
                    "manifest_path": str(manifest_path),
                    "source_record_ids": [str(value) for value in batch.get("source_record_ids", [])],
                }
            )
        _write_json(ledger_path, {"run_id": batch_run_id, "batches": batches_payload})
        _write_json(
            run_dir / "summary.json",
            {
                "run_id": batch_run_id,
                "all_passed": bool(batch_run.get("all_passed", True)),
                "ledger_path": str(ledger_path),
            },
        )


def _write_download_run(*, output_dir: Path, scenario: dict) -> None:
    run_id = str(scenario.get("run_id") or "upstream-run")
    run_dir = output_dir / "runs" / run_id
    manifest_dir = output_dir / "manifests"
    run_dir.mkdir(parents=True, exist_ok=True)
    manifest_dir.mkdir(parents=True, exist_ok=True)
    records = _materialize_records(output_dir=output_dir, records=scenario.get("records") or [])
    manifest_path = manifest_dir / f"download_{run_id}.jsonl"
    _write_jsonl(manifest_path, records)
    status_counts = Counter(str(record["status"]) for record in records)
    summary = {
        "run_id": run_id,
        "mode": str(scenario.get("mode") or "download"),
        "filtered_rows": len(records),
        "filtered_override_count": sum(
            1 for record in records if record.get("original_url") != record.get("effective_url")
        ),
        "status_counts": dict(status_counts),
        "manifest_path": str(manifest_path),
        "workbook_sha256": "upstream-eval",
    }
    summary.update(scenario.get("summary_overrides") or {})
    _write_json(run_dir / "summary.json", summary)


def _materialize_records(*, output_dir: Path, records: list[dict]) -> list[dict]:
    materialized = []
    for definition in records:
        record = {
            "source_record_id": str(definition["source_record_id"]),
            "sheet": str(definition.get("sheet") or "Ingest_Checklist"),
            "excel_row": int(definition.get("excel_row", 5)),
            "title": str(definition.get("title") or definition["source_record_id"]),
            "original_url": str(definition["original_url"]),
            "effective_url": str(definition.get("effective_url") or definition["original_url"]),
            "normalized_url": str(
                definition.get("normalized_url")
                or normalize_url(str(definition.get("effective_url") or definition["original_url"]))
            ),
            "status": str(definition["status"]),
            "http_status": definition.get("http_status"),
            "fetch_timestamp": definition.get("fetch_timestamp"),
            "duplicate_of": None,
            "validation": {
                "mode": str(definition.get("validation_mode") or "download"),
                "passed": bool(definition.get("validation_passed", True)),
            },
            "metadata": dict(definition.get("metadata") or {}),
        }
        duplicate_of = definition.get("duplicate_of")
        if duplicate_of:
            record["duplicate_of"] = str(output_dir / str(duplicate_of))
        artifact = definition.get("artifact")
        artifact_path = definition.get("artifact_path")
        if artifact_path:
            resolved_artifact_path = output_dir / str(artifact_path)
            record["artifact_path"] = str(resolved_artifact_path)
            if artifact:
                resolved_artifact_path.parent.mkdir(parents=True, exist_ok=True)
                body = _artifact_bytes(artifact)
                resolved_artifact_path.write_bytes(body)
                record["artifact_sha256"] = hashlib.sha256(body).hexdigest()
                record["artifact_byte_size"] = len(body)
            else:
                record["artifact_sha256"] = definition.get("artifact_sha256")
                record["artifact_byte_size"] = definition.get("artifact_byte_size")
            if resolved_artifact_path.exists() and (
                record["artifact_sha256"] is None or record["artifact_byte_size"] is None
            ):
                body = resolved_artifact_path.read_bytes()
                record["artifact_sha256"] = hashlib.sha256(body).hexdigest()
                record["artifact_byte_size"] = len(body)
        else:
            record["artifact_path"] = None
            record["artifact_sha256"] = None
            record["artifact_byte_size"] = None
        if definition.get("content_type"):
            record["content_type"] = str(definition["content_type"])
        if definition.get("final_url"):
            record["final_url"] = str(definition["final_url"])
        materialized.append(record)
    return materialized


def _artifact_bytes(spec: dict) -> bytes:
    artifact_type = str(spec["type"])
    if artifact_type == "html":
        return str(spec["body"]).encode("utf-8")
    if artifact_type == "xml":
        return str(spec["body"]).encode("utf-8")
    if artifact_type == "text":
        return str(spec["body"]).encode("utf-8")
    if artifact_type == "pdf":
        if spec.get("pages"):
            return _pdf_bytes(
                [[str(line) for line in page] for page in spec.get("pages", [])]
            )
        return _pdf_bytes([[str(line) for line in spec.get("lines", [])]])
    if artifact_type == "docx":
        return _docx_bytes([str(value) for value in spec.get("paragraphs", [])])
    raise ValueError(f"Unsupported artifact type: {artifact_type}")


def _write_workbook(path: Path, workbook_spec: dict) -> None:
    workbook = Workbook()
    ingest_sheet = workbook.active
    ingest_sheet.title = "Ingest_Checklist"
    forest_plan_sheet = workbook.create_sheet("R1_Forest_Plans")
    exclusions_sheet = workbook.create_sheet("Scope_Exclusions")

    ingest_headers = [
        "ID",
        "Document / Source",
        "Official_Link",
        "Ingest_Status",
        "Scope",
        "Layer",
        "Issuer",
        "Document_Type",
        "Applies_To",
        "Trigger / When to Apply",
        "Review_Engine_Checks",
        "Currentness / Notes",
    ]
    forest_headers = [
        "Unit / Overlay",
        "Current Document or Source",
        "Official_Link",
        "Status in Apr 2026",
        "Applies_When",
        "Review_Engine_Checks",
        "Notes",
    ]
    exclusion_headers = ["Link"]

    _write_sheet_headers(ingest_sheet, ingest_headers)
    _write_sheet_headers(forest_plan_sheet, forest_headers)
    _write_sheet_headers(exclusions_sheet, exclusion_headers)

    for row_index, row in enumerate(workbook_spec.get("ingest_rows", []), start=5):
        ingest_sheet.cell(row_index, 1).value = row["source_record_id"]
        ingest_sheet.cell(row_index, 2).value = row["title"]
        ingest_sheet.cell(row_index, 3).value = row["url"]
        ingest_sheet.cell(row_index, 4).value = row.get("ingest_status", "Ready")
        ingest_sheet.cell(row_index, 5).value = row.get("scope", "Baseline")
        ingest_sheet.cell(row_index, 6).value = row.get("layer", "Federal")
        ingest_sheet.cell(row_index, 7).value = row.get("issuer", "U.S. Congress")
        ingest_sheet.cell(row_index, 8).value = row.get("document_type", "public law")
        ingest_sheet.cell(row_index, 9).value = row.get("applies_to", "Region 1")
        ingest_sheet.cell(row_index, 10).value = row.get("trigger", "Always")
        ingest_sheet.cell(row_index, 11).value = row.get("review_engine_checks", "upstream eval")
        ingest_sheet.cell(row_index, 12).value = row.get("currentness_notes", "fixture")

    for row_index, row in enumerate(workbook_spec.get("forest_plan_rows", []), start=5):
        forest_plan_sheet.cell(row_index, 1).value = row["unit_or_overlay"]
        forest_plan_sheet.cell(row_index, 2).value = row["title"]
        forest_plan_sheet.cell(row_index, 3).value = row["url"]
        forest_plan_sheet.cell(row_index, 4).value = row.get("status", "Current")
        forest_plan_sheet.cell(row_index, 5).value = row.get("applies_when", "Always")
        forest_plan_sheet.cell(row_index, 6).value = row.get("review_engine_checks", "upstream eval")
        forest_plan_sheet.cell(row_index, 7).value = row.get("notes", "fixture")

    for row_index, url in enumerate(workbook_spec.get("excluded_urls", []), start=5):
        exclusions_sheet.cell(row_index, 1).value = url

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _scenario_workbook_path(temp_dir: Path, scenario: dict) -> Path:
    workbook_path = scenario.get("workbook_path")
    if workbook_path is not None:
        return _resolve_case_path(REPO_ROOT, workbook_path)
    workbook_spec = scenario.get("workbook")
    if not isinstance(workbook_spec, dict):
        raise ValueError("Upstream evaluation scenario must provide workbook or workbook_path")
    path = temp_dir / "workbook.xlsx"
    _write_workbook(path, workbook_spec)
    return path


def _write_sheet_headers(sheet, headers: list[str]) -> None:  # noqa: ANN001
    for row_index in range(1, 4):
        sheet.cell(row_index, 1).value = None
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index).value = header


def _pdf_bytes(pages: list[list[str]]) -> bytes:
    if not pages:
        pages = [[]]

    objects: list[bytes] = []
    page_object_ids: list[int] = []
    next_object_id = 1

    catalog_id = next_object_id
    next_object_id += 1
    pages_id = next_object_id
    next_object_id += 1
    font_id = next_object_id
    next_object_id += 1

    objects.append(b"")
    objects.append(b"")
    objects.append(b"")

    for page_lines in pages:
        content_lines = ["BT", "/F1 12 Tf", "72 720 Td"]
        for index, line in enumerate(page_lines):
            escaped = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            if index:
                content_lines.append("0 -16 Td")
            content_lines.append(f"({escaped}) Tj")
        content_lines.append("ET")
        stream = "\n".join(content_lines).encode("utf-8")

        page_id = next_object_id
        content_id = next_object_id + 1
        next_object_id += 2
        page_object_ids.append(page_id)
        objects.append(
            (
                f"{page_id} 0 obj << /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 612 792] "
                f"/Contents {content_id} 0 R /Resources << /Font << /F1 {font_id} 0 R >> >> >> endobj\n"
            ).encode("ascii")
        )
        objects.append(
            f"{content_id} 0 obj << /Length {len(stream)} >> stream\n".encode("ascii")
            + stream
            + b"\nendstream\nendobj\n"
        )

    objects[catalog_id - 1] = f"{catalog_id} 0 obj << /Type /Catalog /Pages {pages_id} 0 R >> endobj\n".encode("ascii")
    kids = " ".join(f"{page_id} 0 R" for page_id in page_object_ids)
    objects[pages_id - 1] = (
        f"{pages_id} 0 obj << /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >> endobj\n"
    ).encode("ascii")
    objects[font_id - 1] = (
        f"{font_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n"
    ).encode("ascii")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_start = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        (
            f"trailer << /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n"
            f"startxref\n{xref_start}\n%%EOF\n"
        ).encode("ascii")
    )
    return bytes(pdf)


def _docx_bytes(paragraphs: list[str]) -> bytes:
    document_body = "".join(
        f"<w:p><w:r><w:t>{_xml_escape(text)}</w:t></w:r></w:p>" for text in paragraphs
    )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{document_body}</w:body>"
        "</w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        "</Relationships>"
    )
    with TemporaryDirectory() as tmp:
        archive_path = Path(tmp) / "fixture.docx"
        with ZipFile(archive_path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", content_types)
            archive.writestr("_rels/.rels", rels)
            archive.writestr("word/document.xml", document_xml)
        return archive_path.read_bytes()


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _apply_extraction_mutations(
    *,
    extraction_manifest_path: Path,
    chunks_path: Path,
    mutations: list[dict],
) -> None:
    if not mutations:
        return
    manifest_records = _read_jsonl(extraction_manifest_path)
    chunks = _read_jsonl(chunks_path)
    manifest_by_source = {
        str(record.get("source_record_id") or ""): record for record in manifest_records
    }
    chunks_by_source: dict[str, list[dict]] = {}
    for chunk in chunks:
        source_record_id = str(chunk.get("source_record_id") or "")
        chunks_by_source.setdefault(source_record_id, []).append(chunk)

    for mutation in mutations:
        source_record_id = str(
            mutation.get("source_record_id")
            or manifest_records[0].get("source_record_id")
            or ""
        )
        record = manifest_by_source[source_record_id]
        text_path = Path(record["text_path"])
        action = str(mutation["action"])
        if action == "overwrite_extracted_text":
            text_path.write_text(str(mutation["text"]), encoding="utf-8")
        elif action == "append_extracted_text":
            text_path.write_text(
                text_path.read_text(encoding="utf-8") + str(mutation["text"]),
                encoding="utf-8",
            )
        elif action == "overwrite_first_chunk_text":
            source_chunks = chunks_by_source.get(source_record_id) or []
            if not source_chunks:
                raise ValueError(f"No chunks available for mutation source_record_id={source_record_id}")
            source_chunks[0]["text"] = str(mutation["text"])
        elif action == "mark_reused_existing":
            parser_metadata = dict(record.get("parser_metadata") or {})
            parser_metadata["reused_existing"] = True
            record["parser_metadata"] = parser_metadata
        else:
            raise ValueError(f"Unsupported extraction mutation action: {action}")

    _write_jsonl(extraction_manifest_path, manifest_records)
    _write_jsonl(chunks_path, chunks)


def _preflight_result_from_spec(spec: dict) -> PreflightFetchResult:
    return PreflightFetchResult(
        status=str(spec["status"]),
        http_status=int(spec["http_status"]) if spec.get("http_status") is not None else None,
        final_url=str(spec.get("final_url") or ""),
        redirect_chain=[str(value) for value in spec.get("redirect_chain", [])],
        content_type=str(spec.get("content_type") or ""),
        content_length=int(spec["content_length"]) if spec.get("content_length") is not None else None,
        method=str(spec.get("method") or "HEAD"),
        failure=spec.get("failure"),
        validation={
            "mode": "preflight",
            "passed": bool(spec.get("validation_passed", True)),
            "reason": spec.get("validation_reason"),
        },
        attempt_count=int(spec.get("attempt_count", 1)),
    )


def _default_preflight_ok(url: str) -> PreflightFetchResult:
    return PreflightFetchResult(
        status="preflight_ok",
        http_status=200,
        final_url=url,
        redirect_chain=[],
        content_type="text/html",
        content_length=4096,
        method="HEAD",
        failure=None,
        validation={"mode": "preflight", "passed": True, "reason": None},
        attempt_count=1,
    )


def _check_outcome(report: dict, name: str) -> bool | None:
    for check in report.get("checks", []):
        if check.get("name") == name:
            return bool(check.get("passed"))
    return None


def _failed_check_names(report: dict) -> list[str]:
    return [
        str(check.get("name"))
        for check in report.get("checks", [])
        if not check.get("passed")
    ]


def _markdown_report(summary: dict) -> str:
    lines = [
        "# Upstream Evaluation Report",
        "",
        f"- Passed: `{summary['passed']}`",
        f"- Contract checks passed: `{sum(1 for check in summary['contract_checks'] if check['passed'])}/{len(summary['contract_checks'])}`",
        f"- Cases matched expectation: `{summary['matched_case_count']}/{summary['case_count']}`",
        "",
        "## Lanes",
        "",
        "| Lane | Status | Direct Eval Categories | Failing Cases |",
        "| --- | --- | ---: | --- |",
    ]
    for lane_summary in summary["lane_summaries"]:
        failing_cases = ", ".join(lane_summary["failing_case_ids"]) or "none"
        lines.append(
            f"| `{lane_summary['lane_id']}` | `{lane_summary['status']}` | "
            f"{lane_summary['direct_eval_present_category_count']}/{lane_summary['category_count']} | "
            f"{failing_cases} |"
        )
    lines.extend(["", "## Failures", ""])
    if summary["failed_case_ids"]:
        for case_id in summary["failed_case_ids"]:
            lines.append(f"- `{case_id}`")
    else:
        lines.append("- none")
    return "\n".join(lines) + "\n"


def _read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line]


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(record, sort_keys=True) + "\n" for record in records),
        encoding="utf-8",
    )


def _resolve_path(value: Path | str | None) -> Path:
    if value is None:
        raise ValueError("Expected path value")
    path = Path(value)
    if path.is_absolute():
        return path
    return (Path.cwd() / path).resolve()


def _resolve_case_path(base_dir: Path, value: Path | str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def _normalized_case_definitions(*, manifest_dir: Path, case_definitions: list[dict]) -> list[dict]:
    normalized = []
    for case_definition in case_definitions:
        fixture_path = case_definition.get("fixture_path")
        if fixture_path is None:
            normalized.append(dict(case_definition))
            continue
        normalized.append(
            {
                **case_definition,
                "fixture_path": str(_resolve_case_path(manifest_dir, fixture_path)),
            }
        )
    return normalized


def _is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
    except ValueError:
        return False
    return True


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _upstream_eval_config(*, loader_contract: str = LEGACY_WORKBOOK_LOADER_CONTRACT):
    config = load_config(REPO_ROOT / DEFAULT_CONFIG_PATH)
    return replace(
        config,
        workbook=replace(
            config.workbook,
            loader_contract=loader_contract,
            overrides_path=None,
        ),
    )
