from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urlsplit
import csv

from openpyxl import load_workbook

from .config import (
    LEGACY_WORKBOOK_LOADER_CONTRACT,
    SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT,
    WorkbookConfig,
)
from .overrides import (
    apply_url_overrides,
    load_url_overrides,
    validate_overrides_do_not_target_exclusions,
)
from .records import WorkbookSource, normalize_url, slugify


INGEST_LINK_COLUMN = "Official_Link"
FOREST_PLAN_LINK_COLUMN = "Official_Link"
EXCLUSION_LINK_COLUMN = "Link"
R1_FOREST_PLAN_DOCUMENT_REGISTER_SHEET = "R1_Forest_Plan_Document_Register"
R1_FOREST_PLAN_REGISTER_SOURCE_DELTA_STATUS = "source_delta_required"
R1_FOREST_PLAN_REGISTER_CATALOG_STATUS = "catalog_confirmed"
R1_FOREST_PLAN_REGISTER_GAP_STATUS = "official_source_gap_documented"
R1_FOREST_PLAN_REGISTER_ALLOWED_STATUSES = {
    R1_FOREST_PLAN_REGISTER_SOURCE_DELTA_STATUS,
    R1_FOREST_PLAN_REGISTER_CATALOG_STATUS,
    R1_FOREST_PLAN_REGISTER_GAP_STATUS,
}
R1_FOREST_PLAN_REGISTER_COLUMNS = (
    "proposed_source_record_id",
    "forest_unit_id",
    "forest_unit_name",
    "document_role",
    "document_title",
    "official_link",
    "existing_source_record_id",
    "readiness_tier",
    "draft_status",
    "required_for",
    "notes",
)


@dataclass(frozen=True)
class R1ForestPlanDocumentRegister:
    path: Path
    rows: list[dict[str, str]]
    source_delta_sources: list[WorkbookSource]
    catalog_confirmed_source_record_ids: list[str]
    gap_source_record_ids: list[str]
    status_counts: dict[str, int]
    forest_unit_ids: list[str]

    def summary(self) -> dict:
        return {
            "schema": "r1-forest-plan-document-register-v0",
            "path": str(self.path),
            "total_rows": len(self.rows),
            "forest_unit_count": len(self.forest_unit_ids),
            "forest_unit_ids": self.forest_unit_ids,
            "status_counts": self.status_counts,
            "source_delta_count": len(self.source_delta_sources),
            "catalog_confirmed_count": len(self.catalog_confirmed_source_record_ids),
            "gap_count": len(self.gap_source_record_ids),
            "source_delta_source_record_ids": [
                source.source_record_id for source in self.source_delta_sources
            ],
            "catalog_confirmed_source_record_ids": self.catalog_confirmed_source_record_ids,
            "skipped_gap_source_record_ids": self.gap_source_record_ids,
        }


def _headers_by_name(sheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(header_row, column).value
        if value not in (None, ""):
            headers[str(value)] = column
    return headers


def _row_has_data(sheet, row: int) -> bool:
    return any(sheet.cell(row, column).value not in (None, "") for column in range(1, sheet.max_column + 1))


def _cell(sheet, row: int, headers: dict[str, int], name: str) -> str | None:
    column = headers.get(name)
    if column is None:
        return None
    value = sheet.cell(row, column).value
    if value in (None, ""):
        return None
    return str(value).strip()


def load_excluded_urls(workbook_path: Path, config: WorkbookConfig) -> set[str]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if config.exclusion_sheet not in workbook.sheetnames:
        return set()

    sheet = workbook[config.exclusion_sheet]
    headers = _headers_by_name(sheet, config.header_row)
    link_column = headers.get(EXCLUSION_LINK_COLUMN)
    if link_column is None:
        return set()

    urls: set[str] = set()
    for row in range(config.header_row + 1, sheet.max_row + 1):
        value = sheet.cell(row, link_column).value
        if value not in (None, ""):
            urls.add(normalize_url(str(value)))
    return urls


def load_canonical_sources(workbook_path: Path, config: WorkbookConfig) -> list[WorkbookSource]:
    if config.loader_contract == LEGACY_WORKBOOK_LOADER_CONTRACT:
        sources = load_legacy_canonical_sources(workbook_path, config)
    elif config.loader_contract == SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT:
        from .source_register import load_source_register_workbook_sources

        sources = load_source_register_workbook_sources(workbook_path)
    else:
        raise ValueError(f"Unsupported workbook loader contract: {config.loader_contract!r}")

    if config.loader_contract == LEGACY_WORKBOOK_LOADER_CONTRACT:
        overrides = load_url_overrides(config.overrides_path)
        sources = apply_url_overrides(sources, overrides)
    validate_overrides_do_not_target_exclusions(sources, load_excluded_urls(workbook_path, config))
    return sources


def load_legacy_canonical_sources(workbook_path: Path, config: WorkbookConfig) -> list[WorkbookSource]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    sources: list[WorkbookSource] = []

    for sheet_name in config.canonical_sheets:
        sheet = workbook[sheet_name]
        headers = _headers_by_name(sheet, config.header_row)
        if sheet_name == "Ingest_Checklist":
            sources.extend(_load_ingest_checklist(sheet, headers, config.header_row))
        elif sheet_name == "R1_Forest_Plans":
            sources.extend(_load_forest_plans(sheet, headers, config.header_row))
        else:
            raise ValueError(f"Unsupported canonical sheet: {sheet_name}")
    return sources


def merge_supplemental_sources(
    workbook_sources: list[WorkbookSource],
    supplemental_sources: Iterable[WorkbookSource] | None = None,
) -> list[WorkbookSource]:
    sources = list(workbook_sources)
    supplemental = list(supplemental_sources or [])
    duplicate_ids = _duplicate_values(
        [source.source_record_id for source in [*sources, *supplemental]]
    )
    if duplicate_ids:
        raise ValueError(f"Supplemental sources duplicate existing source IDs: {duplicate_ids}")
    return [*sources, *supplemental]


def ensure_supplemental_sources_allowed(
    config: WorkbookConfig,
    supplemental_sources: Iterable[WorkbookSource] | None = None,
    source_delta_input: dict | None = None,
) -> None:
    if config.loader_contract != SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT:
        return
    if list(supplemental_sources or []) or source_delta_input is not None:
        raise ValueError(
            "Supplemental source-delta rows are not allowed when "
            "loader_contract='source_register_v1'; the canonical source register "
            "must be the sole active source ledger."
        )


def load_r1_forest_plan_document_register(
    register_path: Path | str,
) -> R1ForestPlanDocumentRegister:
    path = Path(register_path)
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        missing_columns = [
            column for column in R1_FOREST_PLAN_REGISTER_COLUMNS if column not in (reader.fieldnames or [])
        ]
        if missing_columns:
            raise ValueError(f"Missing required R1 forest-plan register columns: {missing_columns}")
        rows = [
            _normalize_register_row(row, csv_row_number=index)
            for index, row in enumerate(reader, start=2)
            if any((value or "").strip() for value in row.values())
        ]

    duplicate_ids = _duplicate_values([row["proposed_source_record_id"] for row in rows])
    if duplicate_ids:
        raise ValueError(f"Duplicate R1 forest-plan register source IDs: {duplicate_ids}")

    invalid_statuses = sorted(
        {
            row["draft_status"]
            for row in rows
            if row["draft_status"] not in R1_FOREST_PLAN_REGISTER_ALLOWED_STATUSES
        }
    )
    if invalid_statuses:
        raise ValueError(f"Unsupported R1 forest-plan register statuses: {invalid_statuses}")

    bad_links = [
        row["proposed_source_record_id"]
        for row in rows
        if urlsplit(row["official_link"]).scheme not in {"http", "https"}
    ]
    if bad_links:
        raise ValueError(f"R1 forest-plan register rows without HTTP(S) official links: {bad_links}")

    source_delta_rows = [
        row for row in rows if row["draft_status"] == R1_FOREST_PLAN_REGISTER_SOURCE_DELTA_STATUS
    ]
    catalog_confirmed_ids = [
        row["proposed_source_record_id"]
        for row in rows
        if row["draft_status"] == R1_FOREST_PLAN_REGISTER_CATALOG_STATUS
    ]
    gap_ids = [
        row["proposed_source_record_id"]
        for row in rows
        if row["draft_status"] == R1_FOREST_PLAN_REGISTER_GAP_STATUS
    ]
    status_counts = Counter(row["draft_status"] for row in rows)
    return R1ForestPlanDocumentRegister(
        path=path,
        rows=rows,
        source_delta_sources=[_register_row_to_workbook_source(row, path) for row in source_delta_rows],
        catalog_confirmed_source_record_ids=catalog_confirmed_ids,
        gap_source_record_ids=gap_ids,
        status_counts=dict(status_counts),
        forest_unit_ids=sorted({row["forest_unit_id"] for row in rows}),
    )


def _load_ingest_checklist(sheet, headers: dict[str, int], header_row: int) -> Iterable[WorkbookSource]:
    required = ["ID", "Document / Source", INGEST_LINK_COLUMN]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing required Ingest_Checklist headers: {missing}")

    for row in range(header_row + 1, sheet.max_row + 1):
        if not _row_has_data(sheet, row):
            continue
        source_id = _cell(sheet, row, headers, "ID")
        title = _cell(sheet, row, headers, "Document / Source")
        original_url = _cell(sheet, row, headers, INGEST_LINK_COLUMN)
        if not source_id or not title or not original_url:
            raise ValueError(f"Ingest_Checklist row {row} is missing ID, title, or URL")
        yield WorkbookSource(
            source_record_id=source_id,
            sheet=sheet.title,
            excel_row=row,
            source_id=source_id,
            title=title,
            original_url=original_url,
            effective_url=original_url,
            normalized_url=normalize_url(original_url),
            metadata={
                "ingest_status": _cell(sheet, row, headers, "Ingest_Status"),
                "scope": _cell(sheet, row, headers, "Scope"),
                "layer": _cell(sheet, row, headers, "Layer"),
                "issuer": _cell(sheet, row, headers, "Issuer"),
                "document_type": _cell(sheet, row, headers, "Document_Type"),
                "applies_to": _cell(sheet, row, headers, "Applies_To"),
                "trigger": _cell(sheet, row, headers, "Trigger / When to Apply"),
                "review_engine_checks": _cell(sheet, row, headers, "Review_Engine_Checks"),
                "currentness_notes": _cell(sheet, row, headers, "Currentness / Notes"),
            },
        )


def _load_forest_plans(sheet, headers: dict[str, int], header_row: int) -> Iterable[WorkbookSource]:
    required = ["Unit / Overlay", "Current Document or Source", FOREST_PLAN_LINK_COLUMN]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing required R1_Forest_Plans headers: {missing}")

    seen_unit_rows: dict[str, int] = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        if not _row_has_data(sheet, row):
            continue
        unit = _cell(sheet, row, headers, "Unit / Overlay")
        title = _cell(sheet, row, headers, "Current Document or Source")
        original_url = _cell(sheet, row, headers, FOREST_PLAN_LINK_COLUMN)
        if not unit or not title or not original_url:
            raise ValueError(f"R1_Forest_Plans row {row} is missing unit, title, or URL")
        unit_slug = slugify(unit, max_length=28)
        seen_unit_rows[unit_slug] = seen_unit_rows.get(unit_slug, 0) + 1
        source_id = f"R1PLAN-{unit_slug}-{seen_unit_rows[unit_slug]:02d}"
        yield WorkbookSource(
            source_record_id=source_id,
            sheet=sheet.title,
            excel_row=row,
            source_id=source_id,
            title=title,
            original_url=original_url,
            effective_url=original_url,
            normalized_url=normalize_url(original_url),
            metadata={
                "unit_or_overlay": unit,
                "status_in_apr_2026": _cell(sheet, row, headers, "Status in Apr 2026"),
                "applies_when": _cell(sheet, row, headers, "Applies_When"),
                "review_engine_checks": _cell(sheet, row, headers, "Review_Engine_Checks"),
                "currentness_notes": _cell(sheet, row, headers, "Notes"),
            },
        )


def _normalize_register_row(row: dict[str, str | None], *, csv_row_number: int) -> dict[str, str]:
    normalized = {
        column: (row.get(column) or "").strip() for column in R1_FOREST_PLAN_REGISTER_COLUMNS
    }
    missing_values = [
        column
        for column in R1_FOREST_PLAN_REGISTER_COLUMNS
        if column != "existing_source_record_id" and not normalized[column]
    ]
    if missing_values:
        raise ValueError(
            f"R1 forest-plan register row {csv_row_number} missing required values: {missing_values}"
        )
    normalized["register_row"] = str(csv_row_number)
    return normalized


def _register_row_to_workbook_source(row: dict[str, str], register_path: Path) -> WorkbookSource:
    source_id = row["proposed_source_record_id"]
    original_url = row["official_link"]
    return WorkbookSource(
        source_record_id=source_id,
        sheet=R1_FOREST_PLAN_DOCUMENT_REGISTER_SHEET,
        excel_row=int(row["register_row"]),
        source_id=source_id,
        title=row["document_title"],
        original_url=original_url,
        effective_url=original_url,
        normalized_url=normalize_url(original_url),
        metadata={
            "source_input": "r1_forest_plan_document_register",
            "register_path": str(register_path),
            "register_row": row["register_row"],
            "forest_unit_id": row["forest_unit_id"],
            "forest_unit_name": row["forest_unit_name"],
            "unit_or_overlay": row["forest_unit_name"],
            "document_role": row["document_role"],
            "document_type": row["document_role"],
            "existing_source_record_id": row["existing_source_record_id"],
            "readiness_tier": row["readiness_tier"],
            "draft_status": row["draft_status"],
            "review_engine_checks": row["required_for"],
            "required_for": row["required_for"],
            "currentness_notes": row["notes"],
            "notes": row["notes"],
        },
    )


def _duplicate_values(values: Iterable[str]) -> list[str]:
    counts = Counter(values)
    return sorted(value for value, count in counts.items() if count > 1)
