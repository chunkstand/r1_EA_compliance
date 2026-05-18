from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from importlib import import_module
from pathlib import Path
from urllib.parse import urlsplit
import json
import re

from openpyxl import load_workbook

from .config import DEFAULT_CONFIG_PATH, load_config
from .records import WorkbookSource, normalize_url, sha256_file, slugify


DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH = Path("config/source_register_sheet_contract_v1.json")
DEFAULT_SOURCE_REGISTER_SCHEMA_PATH = Path("config/source_register_schema_v1.json")
DEFAULT_SOURCE_REGISTER_VOCABULARIES_PATH = Path("config/source_register_vocabularies_v1.json")
DEFAULT_SOURCE_REGISTER_ROW_STATES_PATH = Path("config/source_register_row_states_v1.json")
DEFAULT_DIRECT_FILE_READINESS_CONTRACT_PATH = Path("config/direct_file_readiness_contract_v1.json")
DEFAULT_PARSER_ADMISSION_CONTRACT_PATH = Path("config/parser_admission_contract_v1.json")
DEFAULT_CITATION_ALIAS_REGISTER_PATH = Path("config/citation_alias_register_v1.json")
DEFAULT_JURISDICTION_SCOPE_REGISTER_PATH = Path("config/jurisdiction_scope_register_v1.json")


@dataclass(frozen=True)
class CanonicalSourceRow:
    source_record_id: str
    sheet: str
    excel_row: int
    title: str
    original_url: str
    normalized_url: str
    authority_tier: str
    sub_tier: str | None
    jurisdiction_or_unit: str | None
    resource_area: str | None
    document_type: str | None
    citation_or_code: str | None
    issuing_entity: str | None
    issue_or_effective_date: str | None
    currentness_status: str | None
    applicability_or_trigger: str | None
    url_class: str | None
    validation_status: str | None
    row_state: str
    ea_system_applicability_status: str | None
    criticality_determination: str | None
    applicability_scope: str | None
    ingest_action: str | None
    authority_document_id: str
    authority_document_class_id: str
    authority_section_id: str | None
    jurisdiction_scope_id: str
    source_authority_link_id: str
    direct_file_readiness_class: str
    parser_route_id: str
    parser_admission_class: str
    expected_parser: str
    metadata: dict[str, str | None]

    def to_workbook_source(self) -> WorkbookSource:
        metadata = {
            "loader_contract": "source_register_v1",
            "source_contract_version": "source-register-sheet-contract-v1",
            "row_state": self.row_state,
            "authority_tier": self.authority_tier,
            "sub_tier": self.sub_tier,
            "jurisdiction_or_unit": self.jurisdiction_or_unit,
            "resource_area": self.resource_area,
            "document_type": self.document_type,
            "citation_or_code": self.citation_or_code,
            "issuer": self.issuing_entity,
            "issue_or_effective_date": self.issue_or_effective_date,
            "currentness_status": self.currentness_status,
            "currentness_notes": self.currentness_status,
            "applies_to": self.ea_system_applicability_status,
            "trigger": self.applicability_or_trigger,
            "review_engine_checks": self.applicability_scope,
            "url_class": self.url_class,
            "validation_status": self.validation_status,
            "criticality_determination": self.criticality_determination,
            "applicability_scope": self.applicability_scope,
            "ingest_action": self.ingest_action,
            "authority_document_id": self.authority_document_id,
            "authority_document_class_id": self.authority_document_class_id,
            "authority_section_id": self.authority_section_id,
            "jurisdiction_scope_id": self.jurisdiction_scope_id,
            "source_authority_link_id": self.source_authority_link_id,
            "direct_file_readiness_class": self.direct_file_readiness_class,
            "parser_route_id": self.parser_route_id,
            "parser_admission_class": self.parser_admission_class,
            "expected_parser": self.expected_parser,
            **self.metadata,
        }
        return WorkbookSource(
            source_record_id=self.source_record_id,
            sheet=self.sheet,
            excel_row=self.excel_row,
            source_id=self.source_record_id,
            title=self.title,
            original_url=self.original_url,
            effective_url=self.original_url,
            normalized_url=self.normalized_url,
            metadata=metadata,
        )


def load_source_register_contract(path: Path | str = DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "source-register-sheet-contract-v1":
        raise ValueError(
            "Unsupported source register sheet contract schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_source_register_schema(path: Path | str = DEFAULT_SOURCE_REGISTER_SCHEMA_PATH) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "source-register-schema-v1":
        raise ValueError(
            "Unsupported source register schema schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_source_register_vocabularies(
    path: Path | str = DEFAULT_SOURCE_REGISTER_VOCABULARIES_PATH,
) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "source-register-vocabularies-v1":
        raise ValueError(
            "Unsupported source register vocabularies schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_source_register_row_states(path: Path | str = DEFAULT_SOURCE_REGISTER_ROW_STATES_PATH) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "source-register-row-states-v1":
        raise ValueError(
            "Unsupported source register row-states schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_direct_file_readiness_contract(
    path: Path | str = DEFAULT_DIRECT_FILE_READINESS_CONTRACT_PATH,
) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "direct-file-readiness-contract-v1":
        raise ValueError(
            "Unsupported direct-file-readiness contract schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_parser_admission_contract(
    path: Path | str = DEFAULT_PARSER_ADMISSION_CONTRACT_PATH,
) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "parser-admission-contract-v1":
        raise ValueError(
            "Unsupported parser-admission contract schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_citation_alias_register(
    path: Path | str = DEFAULT_CITATION_ALIAS_REGISTER_PATH,
) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "citation-alias-register-v1":
        raise ValueError(
            "Unsupported citation-alias register schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def load_jurisdiction_scope_register(
    path: Path | str = DEFAULT_JURISDICTION_SCOPE_REGISTER_PATH,
) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != "jurisdiction-scope-register-v1":
        raise ValueError(
            "Unsupported jurisdiction-scope register schema_version: "
            f"{payload.get('schema_version')!r}"
        )
    return payload


def read_source_register_tables(
    workbook_path: Path | str,
    contract_path: Path | str = DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH,
) -> dict[str, dict[str, object]]:
    contract = load_source_register_contract(contract_path)
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    tables: dict[str, dict[str, object]] = {}
    for spec in contract["sheets"]:
        sheet_name = spec["sheet_name"]
        if sheet_name not in workbook.sheetnames:
            continue
        tables[sheet_name] = _parse_sheet_table(workbook[sheet_name], spec)
    return tables


def load_source_register_rows(
    workbook_path: Path | str,
    *,
    sheet_contract_path: Path | str = DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH,
    schema_path: Path | str = DEFAULT_SOURCE_REGISTER_SCHEMA_PATH,
    vocabularies_path: Path | str = DEFAULT_SOURCE_REGISTER_VOCABULARIES_PATH,
    row_states_path: Path | str = DEFAULT_SOURCE_REGISTER_ROW_STATES_PATH,
    direct_file_readiness_path: Path | str = DEFAULT_DIRECT_FILE_READINESS_CONTRACT_PATH,
    parser_admission_path: Path | str = DEFAULT_PARSER_ADMISSION_CONTRACT_PATH,
    alias_register_path: Path | str = DEFAULT_CITATION_ALIAS_REGISTER_PATH,
    jurisdiction_scope_path: Path | str = DEFAULT_JURISDICTION_SCOPE_REGISTER_PATH,
) -> list[CanonicalSourceRow]:
    report = validate_source_register(
        workbook_path,
        mode="schema",
        sheet_contract_path=sheet_contract_path,
        schema_path=schema_path,
        vocabularies_path=vocabularies_path,
        row_states_path=row_states_path,
    )
    if not report["validation_passed"]:
        failed_checks = [check["name"] for check in report["checks"] if not check["passed"]]
        raise ValueError(
            "Canonical source register failed schema validation before loader admission: "
            f"{failed_checks}"
        )

    contract = load_source_register_contract(sheet_contract_path)
    row_states = load_source_register_row_states(row_states_path)
    direct_file_contract = load_direct_file_readiness_contract(direct_file_readiness_path)
    parser_contract = load_parser_admission_contract(parser_admission_path)
    alias_register = load_citation_alias_register(alias_register_path)
    scope_register = load_jurisdiction_scope_register(jurisdiction_scope_path)
    tables = read_source_register_tables(workbook_path, contract_path=sheet_contract_path)
    master_rows = _rows_for_sheet(tables, contract["load_sheet_name"])

    row_state = _row_state_for_sheet(row_states, contract["load_sheet_name"])
    readiness_class = _readiness_class_for_sheet(
        direct_file_contract,
        contract["load_sheet_name"],
    )
    alias_index = _alias_index(alias_register)
    rows: list[CanonicalSourceRow] = []
    seen_link_ids: set[str] = set()
    seen_identity_keys: dict[str, str] = {}

    for raw_row in master_rows:
        source_id = str(raw_row["Source_ID"]).strip()
        title = str(raw_row["Document_Title"]).strip()
        original_url = str(raw_row["Source_URL"]).strip()
        route = _resolve_parser_route(raw_row, parser_contract)
        authority_document_id, authority_document_class_id = _resolve_authority_document_identity(
            raw_row,
            alias_register,
            alias_index,
        )
        authority_section_id = _derive_authority_section_id(raw_row, authority_document_id)
        jurisdiction_scope_id = _resolve_jurisdiction_scope_id(raw_row, scope_register)
        source_authority_link_id = f"source_authority_link:{slugify(source_id, max_length=96)}"
        if source_authority_link_id in seen_link_ids:
            raise ValueError(f"Duplicate source-authority link id generated for {source_id}")
        seen_link_ids.add(source_authority_link_id)

        identity_key = _identity_validation_key(raw_row)
        existing_id = seen_identity_keys.get(identity_key)
        if existing_id and existing_id != authority_document_id:
            raise ValueError(
                "Canonical loader identity fork detected for "
                f"{source_id}: {existing_id!r} vs {authority_document_id!r}"
            )
        seen_identity_keys.setdefault(identity_key, authority_document_id)

        rows.append(
            CanonicalSourceRow(
                source_record_id=source_id,
                sheet=contract["load_sheet_name"],
                excel_row=int(raw_row.get("__excel_row__") or 0),
                title=title,
                original_url=original_url,
                normalized_url=normalize_url(original_url),
                authority_tier=_string_or_none(raw_row.get("Authority_Tier")) or "",
                sub_tier=_string_or_none(raw_row.get("Sub_Tier")),
                jurisdiction_or_unit=_string_or_none(raw_row.get("Jurisdiction_or_Unit")),
                resource_area=_string_or_none(raw_row.get("Resource_Area")),
                document_type=_string_or_none(raw_row.get("Document_Type")),
                citation_or_code=_string_or_none(raw_row.get("Citation_or_Code")),
                issuing_entity=_string_or_none(raw_row.get("Issuing_Entity")),
                issue_or_effective_date=_string_or_none(raw_row.get("Issue_or_Effective_Date")),
                currentness_status=_string_or_none(raw_row.get("Currentness_Status")),
                applicability_or_trigger=_string_or_none(raw_row.get("Applicability_or_Trigger")),
                url_class=_string_or_none(raw_row.get("URL_Class")),
                validation_status=_string_or_none(raw_row.get("Validation_Status")),
                row_state=row_state,
                ea_system_applicability_status=_string_or_none(
                    raw_row.get("EA_System_Applicability_Status")
                ),
                criticality_determination=_string_or_none(
                    raw_row.get("Criticality_Determination")
                ),
                applicability_scope=_string_or_none(raw_row.get("Applicability_Scope")),
                ingest_action=_string_or_none(raw_row.get("Ingest_Action")),
                authority_document_id=authority_document_id,
                authority_document_class_id=authority_document_class_id,
                authority_section_id=authority_section_id,
                jurisdiction_scope_id=jurisdiction_scope_id,
                source_authority_link_id=source_authority_link_id,
                direct_file_readiness_class=readiness_class,
                parser_route_id=route["route_id"],
                parser_admission_class=route["admission_class"],
                expected_parser=route["expected_parser"],
                metadata={
                    "link_status": _string_or_none(raw_row.get("Link_Status")),
                    "validation_method": _string_or_none(raw_row.get("Validation_Method")),
                    "docling_instructions": _string_or_none(raw_row.get("Docling_Instructions")),
                    "related_workbook_row": _string_or_none(raw_row.get("Related_Workbook_Row")),
                    "notes": _string_or_none(raw_row.get("Notes")),
                    "source_fingerprint": _string_or_none(raw_row.get("Source_Fingerprint")),
                    "database_load": _string_or_none(raw_row.get("Database_Load")),
                },
            )
        )

    return rows


def load_source_register_workbook_sources(
    workbook_path: Path | str,
) -> list[WorkbookSource]:
    return [row.to_workbook_source() for row in load_source_register_rows(workbook_path)]


def _row_state_for_sheet(row_states: dict, sheet_name: str) -> str:
    for state in row_states["states"]:
        if state.get("sheet_name") == sheet_name:
            return str(state["state_id"])
    raise ValueError(f"No row state configured for source-register sheet {sheet_name!r}")


def _readiness_class_for_sheet(contract: dict, sheet_name: str) -> str:
    for rule in contract["sheet_rules"]:
        if rule.get("sheet_name") == sheet_name:
            return str(rule["readiness_state"])
    raise ValueError(f"No direct-file readiness rule configured for {sheet_name!r}")


def _alias_index(alias_register: dict) -> dict[str, list[tuple[str, str]]]:
    index: dict[str, list[tuple[str, str]]] = {}
    for group in [*alias_register.get("starter_alias_groups", []), *alias_register.get("alias_rows", [])]:
        canonical_id = str(group["canonical_id"])
        canonical_class_id = str(group["canonical_class_id"])
        for alias in group.get("aliases", []):
            text = _normalize_identity_text(alias.get("alias_text"))
            if not text:
                continue
            index.setdefault(text, []).append((canonical_id, canonical_class_id))
    return index


def _resolve_authority_document_identity(
    row: dict[str, object],
    alias_register: dict,
    alias_index: dict[str, list[tuple[str, str]]],
) -> tuple[str, str]:
    title = _string_or_none(row.get("Document_Title"))
    citation = _string_or_none(row.get("Citation_or_Code"))
    candidate_texts = [text for text in (citation, title) if text]
    matches: set[tuple[str, str]] = set()
    for text in candidate_texts:
        matches.update(alias_index.get(_normalize_identity_text(text), []))

    if len({match[0] for match in matches}) > 1:
        raise ValueError(
            "Canonical loader alias resolution is ambiguous for "
            f"{row.get('Source_ID')!r}: {sorted(match[0] for match in matches)}"
        )
    if matches:
        canonical_id, class_id = next(iter(matches))
        return canonical_id, class_id

    blocked_terms = {
        _normalize_identity_text(term)
        for term in alias_register.get("ambiguity_policy", {}).get("blocked_without_context", [])
    }
    context_fields = [
        _string_or_none(row.get("Issuing_Entity")),
        _string_or_none(row.get("Jurisdiction_or_Unit")),
        _string_or_none(row.get("Issue_or_Effective_Date")),
    ]
    if any(_normalize_identity_text(text) in blocked_terms for text in candidate_texts) and not any(
        context_fields
    ):
        raise ValueError(
            "Canonical loader requires more context before resolving blocked alias for "
            f"{row.get('Source_ID')!r}"
        )

    class_id = _infer_authority_document_class_id(row)
    return _derive_generated_authority_document_id(row, class_id), class_id


def _infer_authority_document_class_id(row: dict[str, object]) -> str:
    authority_tier = _normalize_identity_text(row.get("Authority_Tier"))
    document_type = _normalize_identity_text(row.get("Document_Type"))
    title = _normalize_identity_text(row.get("Document_Title"))
    if authority_tier == "forest" and ("forest plan" in document_type or "forest plan" in title):
        return "forest_plan"
    if "forest plan chapter" in document_type or "plan amendment" in document_type:
        return "forest_plan"
    return "authority_document"


def _derive_generated_authority_document_id(row: dict[str, object], class_id: str) -> str:
    parts = [
        _string_or_none(row.get("Authority_Tier")),
        _string_or_none(row.get("Jurisdiction_or_Unit")),
        _string_or_none(row.get("Document_Type")),
        _string_or_none(row.get("Citation_or_Code")),
        _string_or_none(row.get("Document_Title")),
    ]
    stem = slugify("-".join(part for part in parts if part), max_length=120)
    prefix = "forest_plan" if class_id == "forest_plan" else "authority_document"
    return f"{prefix}:{stem}"


def _derive_authority_section_id(
    row: dict[str, object],
    authority_document_id: str,
) -> str | None:
    document_type = _normalize_identity_text(row.get("Document_Type"))
    citation = _string_or_none(row.get("Citation_or_Code"))
    title = _string_or_none(row.get("Document_Title"))
    if citation and any(token in citation.lower() for token in ("§", "chapter", "part", "sec.")):
        return f"{authority_document_id}#section:{slugify(citation, max_length=72)}"
    if "chapter" in document_type and title:
        return f"{authority_document_id}#section:{slugify(title, max_length=72)}"
    return None


def _resolve_jurisdiction_scope_id(row: dict[str, object], scope_register: dict) -> str:
    available_scopes = {str(entry["scope_id"]) for entry in scope_register.get("scopes", [])}
    authority_tier = _string_or_none(row.get("Authority_Tier")) or ""
    scope_by_tier = {
        "Federal": "scope:federal-us",
        "USDA": "scope:usda",
        "USFS": "scope:usfs-national",
        "Region": "scope:usfs-region-1",
        "Forest": "scope:region1-forest-unit",
        "Programmatic": "scope:ea-project-review",
        "State/Partner": "scope:ea-project-review",
        "Superseded": "scope:ea-project-review",
    }
    scope_id = scope_by_tier.get(authority_tier, "scope:ea-project-review")
    if scope_id not in available_scopes:
        raise ValueError(f"Loader-derived scope id {scope_id!r} is not defined in the scope register")
    return scope_id


def _resolve_parser_route(row: dict[str, object], parser_contract: dict) -> dict[str, str]:
    url = _string_or_none(row.get("Source_URL")) or ""
    url_class = _string_or_none(row.get("URL_Class")) or ""
    host = urlsplit(url).netloc.lower()
    path = urlsplit(url).path.lower()
    for route in parser_contract["parser_routes"]:
        match = route.get("match", {})
        if match.get("default"):
            return {
                "route_id": str(route["route_id"]),
                "expected_parser": str(route["expected_parser"]),
                "admission_class": str(route["admission_class"]),
            }
        url_suffixes = [suffix.lower() for suffix in match.get("url_suffixes", [])]
        url_classes = {str(value) for value in match.get("url_classes", [])}
        host_suffixes = [suffix.lower() for suffix in match.get("host_suffixes", [])]
        if url_suffixes and any(path.endswith(suffix) for suffix in url_suffixes):
            if not url_classes or url_class in url_classes:
                return {
                    "route_id": str(route["route_id"]),
                    "expected_parser": str(route["expected_parser"]),
                    "admission_class": str(route["admission_class"]),
                }
        if host_suffixes and any(host == suffix or host.endswith(f".{suffix}") for suffix in host_suffixes):
            return {
                "route_id": str(route["route_id"]),
                "expected_parser": str(route["expected_parser"]),
                "admission_class": str(route["admission_class"]),
            }
    raise ValueError(f"No parser route matched canonical source row {row.get('Source_ID')!r}")


def _identity_validation_key(row: dict[str, object]) -> str:
    parts = (
        _normalize_identity_text(row.get("Document_Title")),
        _normalize_identity_text(row.get("Citation_or_Code")),
        _normalize_identity_text(row.get("Issuing_Entity")),
        _normalize_identity_text(row.get("Jurisdiction_or_Unit")),
    )
    return "|".join(parts)


def _normalize_identity_text(value: object) -> str:
    text = _string_or_none(value)
    if not text:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _string_or_none(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def validate_source_register(
    workbook_path: Path | str,
    *,
    mode: str = "schema",
    sheet_contract_path: Path | str = DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH,
    schema_path: Path | str = DEFAULT_SOURCE_REGISTER_SCHEMA_PATH,
    vocabularies_path: Path | str = DEFAULT_SOURCE_REGISTER_VOCABULARIES_PATH,
    row_states_path: Path | str = DEFAULT_SOURCE_REGISTER_ROW_STATES_PATH,
) -> dict:
    if mode != "schema":
        raise ValueError(f"Unsupported source register validation mode: {mode!r}")

    workbook_path = Path(workbook_path)
    contract = load_source_register_contract(sheet_contract_path)
    schema = load_source_register_schema(schema_path)
    vocabularies = load_source_register_vocabularies(vocabularies_path)
    row_states = load_source_register_row_states(row_states_path)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    tables = read_source_register_tables(workbook_path, contract_path=sheet_contract_path)
    checks: list[dict[str, object]] = []

    _append_check(
        checks,
        name="sheet_roster_matches_contract",
        expected=contract["sheet_order"],
        actual=workbook.sheetnames,
        passed=workbook.sheetnames == contract["sheet_order"],
        details="Workbook sheet order must match the frozen Phase 0 sheet contract.",
    )

    for spec in contract["sheets"]:
        sheet_name = spec["sheet_name"]
        table = tables.get(sheet_name)
        if table is None:
            _append_check(
                checks,
                name=f"{sheet_name}_present",
                expected=True,
                actual=False,
                passed=False,
                details="Required sheet is missing from the canonical source register.",
            )
            continue
        _append_check(
            checks,
            name=f"{sheet_name}_required_columns_present",
            expected=spec["required_columns"],
            actual=table["headers"],
            passed=_contains_all(table["headers"], spec["required_columns"]),
            details="Required columns must exist at the frozen header row.",
        )

    master_rows = _rows_for_sheet(tables, contract["load_sheet_name"])
    queue_rows = _rows_for_sheet(tables, contract["queue_sheet_name"])
    removed_rows = _rows_for_sheet(tables, contract["removed_sheet_name"])

    master_schema = schema["master_sheet"]
    queue_schema = schema["queue_sheet"]
    removed_schema = schema["removed_sheet"]
    vocabulary_values = vocabularies["controlled_values"]

    master_source_ids = [str(row["Source_ID"]).strip() for row in master_rows]
    master_urls = [str(row["Source_URL"]).strip() for row in master_rows]
    master_normalized_urls = [normalize_url(url) for url in master_urls if url]
    blank_master_urls = [row["Source_ID"] for row in master_rows if not str(row["Source_URL"] or "").strip()]
    non_http_master_urls = [
        row["Source_ID"]
        for row in master_rows
        if urlsplit(str(row["Source_URL"] or "").strip()).scheme not in {"http", "https"}
    ]
    missing_master_applicability = [
        row["Source_ID"]
        for row in master_rows
        if not str(row.get("EA_System_Applicability_Status") or "").strip()
    ]
    duplicate_source_ids = _duplicates(master_source_ids)
    duplicate_source_urls = _duplicates(master_normalized_urls)

    _append_check(
        checks,
        name="master_required_nonempty_columns",
        expected=master_schema["required_nonempty_columns"],
        actual=_missing_required_columns(master_rows, master_schema["required_nonempty_columns"]),
        passed=not _missing_required_columns(master_rows, master_schema["required_nonempty_columns"]),
        details="Required master-sheet fields must be populated for every retained load row.",
    )
    _append_check(
        checks,
        name="master_source_id_unique",
        expected=0,
        actual=len(duplicate_source_ids),
        passed=not duplicate_source_ids,
        details="Document_Register_Master Source_ID values must be unique.",
    )
    _append_check(
        checks,
        name="master_source_url_unique",
        expected=0,
        actual=len(duplicate_source_urls),
        passed=not duplicate_source_urls,
        details="Document_Register_Master Source_URL values must be unique after normalization.",
    )
    _append_check(
        checks,
        name="master_source_url_nonempty",
        expected=0,
        actual=len(blank_master_urls),
        passed=not blank_master_urls,
        details="Document_Register_Master rows must not have blank Source_URL values.",
    )
    _append_check(
        checks,
        name="master_source_url_http",
        expected=0,
        actual=len(non_http_master_urls),
        passed=not non_http_master_urls,
        details="Document_Register_Master rows must use HTTP(S) source URLs.",
    )
    _append_check(
        checks,
        name="master_applicability_complete",
        expected=0,
        actual=len(missing_master_applicability),
        passed=not missing_master_applicability,
        details="Document_Register_Master rows must carry EA_System_Applicability_Status.",
    )
    _append_check(
        checks,
        name="master_database_load_values",
        expected=vocabulary_values["master_database_load"],
        actual=sorted({str(row["Database_Load"]).strip() for row in master_rows}),
        passed=_only_allowed_values(
            master_rows,
            "Database_Load",
            vocabulary_values["master_database_load"],
        ),
        details="Retained master rows must stay in the load-ready state.",
    )
    _append_check(
        checks,
        name="master_authority_tier_values",
        expected=vocabulary_values["master_authority_tier"],
        actual=sorted({str(row["Authority_Tier"]).strip() for row in master_rows}),
        passed=_only_allowed_values(
            master_rows,
            "Authority_Tier",
            vocabulary_values["master_authority_tier"],
        ),
        details="Master-sheet authority tiers must stay within the frozen vocabulary.",
    )
    _append_check(
        checks,
        name="master_applicability_status_values",
        expected=vocabulary_values["load_applicability_statuses"],
        actual=sorted({str(row["EA_System_Applicability_Status"]).strip() for row in master_rows}),
        passed=_only_allowed_values(
            master_rows,
            "EA_System_Applicability_Status",
            vocabulary_values["load_applicability_statuses"],
        ),
        details="Master-sheet applicability statuses must stay within the frozen vocabulary.",
    )

    _append_check(
        checks,
        name="queue_required_nonempty_columns",
        expected=queue_schema["required_nonempty_columns"],
        actual=_missing_required_columns(queue_rows, queue_schema["required_nonempty_columns"]),
        passed=not _missing_required_columns(queue_rows, queue_schema["required_nonempty_columns"]),
        details="Queue rows must preserve the fields needed for later direct-file promotion.",
    )
    _append_check(
        checks,
        name="queue_database_load_values",
        expected=vocabulary_values["queue_database_load"],
        actual=sorted({str(row["Database_Load"]).strip() for row in queue_rows}),
        passed=_only_allowed_values(
            queue_rows,
            "Database_Load",
            vocabulary_values["queue_database_load"],
        ),
        details="Direct-file capture queue rows must remain non-load rows.",
    )
    _append_check(
        checks,
        name="queue_authority_tier_values",
        expected=vocabulary_values["queue_authority_tier"],
        actual=sorted({str(row["Authority_Tier"]).strip() for row in queue_rows}),
        passed=_only_allowed_values(
            queue_rows,
            "Authority_Tier",
            vocabulary_values["queue_authority_tier"],
        ),
        details="Queue-sheet authority tiers must stay within the frozen vocabulary.",
    )

    _append_check(
        checks,
        name="removed_required_nonempty_columns",
        expected=removed_schema["required_nonempty_columns"],
        actual=_missing_required_columns(removed_rows, removed_schema["required_nonempty_columns"]),
        passed=not _missing_required_columns(removed_rows, removed_schema["required_nonempty_columns"]),
        details="Removed rows must retain removal rationale and applicability state.",
    )
    _append_check(
        checks,
        name="removed_applicability_status_values",
        expected=vocabulary_values["removed_applicability_statuses"],
        actual=sorted({str(row["EA_System_Applicability_Status"]).strip() for row in removed_rows}),
        passed=_only_allowed_values(
            removed_rows,
            "EA_System_Applicability_Status",
            vocabulary_values["removed_applicability_statuses"],
        ),
        details="Removed rows must remain explicitly non-load audit rows.",
    )

    codex_metrics = _metric_rows_to_map(_rows_for_sheet(tables, "Codex_Load_Index"), "Metric", "Value")
    final_ingest_metrics = _metric_rows_to_map(
        _rows_for_sheet(tables, "Final_Ingest_Certification_2026"),
        "Metric",
        "Value",
    )
    link_summary = _metric_rows_to_map(
        _rows_for_sheet(tables, "Link_Validation_Summary"),
        "Category",
        "Count",
    )
    authority_tier_total = _sum_numeric_field(
        _rows_for_sheet(tables, "Authority_Tier_Index"),
        "Database_Load_Rows",
    )
    applicability_total = _sum_numeric_field(
        _rows_for_sheet(tables, "Applicability_By_Tier"),
        "Row_Count",
    )

    _append_check(
        checks,
        name="codex_master_count_matches_actual",
        expected=len(master_rows),
        actual=_coerce_int(codex_metrics.get("Master database source rows")),
        passed=_coerce_int(codex_metrics.get("Master database source rows")) == len(master_rows),
        details="Codex_Load_Index must match the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="codex_queue_count_matches_actual",
        expected=len(queue_rows),
        actual=_coerce_int(codex_metrics.get("Capture queue rows")),
        passed=_coerce_int(codex_metrics.get("Capture queue rows")) == len(queue_rows),
        details="Codex_Load_Index must match the actual queue-row count.",
    )
    _append_check(
        checks,
        name="codex_removed_count_matches_actual",
        expected=len(removed_rows),
        actual=_coerce_int(codex_metrics.get("Rows removed in this pass")),
        passed=_coerce_int(codex_metrics.get("Rows removed in this pass")) == len(removed_rows),
        details="Codex_Load_Index must match the actual removed-row count.",
    )
    _append_check(
        checks,
        name="final_ingest_master_count_matches_actual",
        expected=len(master_rows),
        actual=_coerce_int(final_ingest_metrics.get("Final retained database-load rows")),
        passed=_coerce_int(final_ingest_metrics.get("Final retained database-load rows")) == len(master_rows),
        details="Final_Ingest_Certification_2026 must match the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="link_summary_master_count_matches_actual",
        expected=len(master_rows),
        actual=_coerce_int(link_summary.get("Working official links in master")),
        passed=_coerce_int(link_summary.get("Working official links in master")) == len(master_rows),
        details="Link_Validation_Summary must match the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="link_summary_queue_count_matches_actual",
        expected=len(queue_rows),
        actual=_coerce_int(link_summary.get("Capture queue non-load rows")),
        passed=_coerce_int(link_summary.get("Capture queue non-load rows")) == len(queue_rows),
        details="Link_Validation_Summary must match the actual queue-row count.",
    )
    _append_check(
        checks,
        name="authority_tier_index_total_matches_actual",
        expected=len(master_rows),
        actual=authority_tier_total,
        passed=authority_tier_total == len(master_rows),
        details="Authority_Tier_Index totals must reconcile to the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="applicability_by_tier_total_matches_actual",
        expected=len(master_rows),
        actual=applicability_total,
        passed=applicability_total == len(master_rows),
        details="Applicability_By_Tier totals must reconcile to the actual retained master-row count.",
    )

    validation_passed = all(bool(check["passed"]) for check in checks)
    state_sheet_names = {
        state["state_id"]: state.get("sheet_name") or state.get("sheet_names")
        for state in row_states["states"]
    }
    return {
        "schema_version": "source-register-validation-v1",
        "mode": mode,
        "workbook_path": str(workbook_path),
        "workbook_sha256": sha256_file(workbook_path),
        "sheet_contract_path": str(Path(sheet_contract_path)),
        "schema_path": str(Path(schema_path)),
        "vocabularies_path": str(Path(vocabularies_path)),
        "row_states_path": str(Path(row_states_path)),
        "sheet_count": len(workbook.sheetnames),
        "load_sheet_name": contract["load_sheet_name"],
        "queue_sheet_name": contract["queue_sheet_name"],
        "removed_sheet_name": contract["removed_sheet_name"],
        "load_row_count": len(master_rows),
        "queue_row_count": len(queue_rows),
        "removed_row_count": len(removed_rows),
        "stale_source_detector_count": sum(
            1
            for row in master_rows
            if str(row.get("EA_System_Applicability_Status") or "").strip()
            == "Applicable - stale-source detector only"
        ),
        "row_state_sheet_names": state_sheet_names,
        "checks": checks,
        "issue_count": sum(1 for check in checks if not check["passed"]),
        "validation_passed": validation_passed,
    }


def build_source_register_diff(
    legacy_workbook_path: Path | str,
    legacy_register_path: Path | str,
    canonical_workbook_path: Path | str,
    *,
    config_path: Path | str = DEFAULT_CONFIG_PATH,
    sheet_contract_path: Path | str = DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH,
) -> dict:
    workbook_module = import_module("usfs_r1_ea_sources.workbook")
    legacy_workbook_path = Path(legacy_workbook_path)
    legacy_register_path = Path(legacy_register_path)
    canonical_workbook_path = Path(canonical_workbook_path)
    config = load_config(config_path)
    legacy_sources = workbook_module.load_legacy_canonical_sources(
        legacy_workbook_path,
        config.workbook,
    )
    register = workbook_module.load_r1_forest_plan_document_register(legacy_register_path)
    tables = read_source_register_tables(
        canonical_workbook_path,
        contract_path=sheet_contract_path,
    )
    master_rows = _rows_for_sheet(tables, "Document_Register_Master")
    queue_rows = _rows_for_sheet(tables, "Direct_File_Capture_Queue")
    removed_rows = _rows_for_sheet(tables, "Removed_Not_Applicable_Final")

    legacy_workbook_ids = {source.source_record_id for source in legacy_sources}
    legacy_source_delta_ids = {source.source_record_id for source in register.source_delta_sources}
    legacy_combined_ids = legacy_workbook_ids | legacy_source_delta_ids
    canonical_ids = {str(row["Source_ID"]).strip() for row in master_rows}
    stale_source_detector_ids = sorted(
        str(row["Source_ID"]).strip()
        for row in master_rows
        if str(row.get("EA_System_Applicability_Status") or "").strip()
        == "Applicable - stale-source detector only"
    )

    queue_reason_counts = Counter(
        str(row.get("Queue_Reason") or "").strip() for row in queue_rows
    )
    applicability_counts = Counter(
        str(row.get("EA_System_Applicability_Status") or "").strip() for row in master_rows
    )

    return {
        "schema_version": "source-register-diff-v1",
        "legacy_workbook_path": str(legacy_workbook_path),
        "legacy_workbook_sha256": sha256_file(legacy_workbook_path),
        "legacy_register_path": str(legacy_register_path),
        "canonical_workbook_path": str(canonical_workbook_path),
        "canonical_workbook_sha256": sha256_file(canonical_workbook_path),
        "legacy_workbook_source_count": len(legacy_workbook_ids),
        "legacy_workbook_unique_url_count": len(
            {source.normalized_url for source in legacy_sources}
        ),
        "legacy_register_source_delta_count": len(legacy_source_delta_ids),
        "legacy_register_catalog_confirmed_count": len(register.catalog_confirmed_source_record_ids),
        "legacy_register_gap_count": len(register.gap_source_record_ids),
        "legacy_runtime_unique_source_count": len(legacy_combined_ids),
        "canonical_master_row_count": len(master_rows),
        "canonical_queue_row_count": len(queue_rows),
        "canonical_removed_row_count": len(removed_rows),
        "canonical_stale_source_detector_count": len(stale_source_detector_ids),
        "canonical_master_unique_url_count": len(
            {normalize_url(str(row["Source_URL"]).strip()) for row in master_rows}
        ),
        "canonical_shared_with_legacy_workbook_count": len(canonical_ids & legacy_workbook_ids),
        "canonical_shared_with_source_delta_count": len(canonical_ids & legacy_source_delta_ids),
        "canonical_only_source_count": len(canonical_ids - legacy_combined_ids),
        "legacy_only_source_count": len(legacy_combined_ids - canonical_ids),
        "canonical_only_source_ids_sample": sorted(canonical_ids - legacy_combined_ids)[:25],
        "legacy_only_source_ids_sample": sorted(legacy_combined_ids - canonical_ids)[:25],
        "canonical_queue_reason_counts": dict(queue_reason_counts),
        "canonical_master_applicability_counts": dict(applicability_counts),
        "canonical_stale_source_detector_ids": stale_source_detector_ids,
    }


def _parse_sheet_table(sheet, spec: dict) -> dict[str, object]:
    header_row = int(spec["header_row"])
    data_start_row = int(spec.get("data_start_row") or (header_row + 1))
    stop_at_blank_row = bool(spec.get("stop_at_blank_row"))

    header_cells = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )
    headers = [str(value).strip() if value not in (None, "") else None for value in header_cells]

    rows = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row,
    ):
        if not _row_has_data(values):
            if stop_at_blank_row:
                break
            continue
        row = {
            header: value
            for header, value in zip(headers, values, strict=False)
            if header not in (None, "")
        }
        row["__excel_row__"] = row_number
        rows.append(row)
    return {
        "headers": [header for header in headers if header not in (None, "")],
        "rows": rows,
        "header_row": header_row,
        "data_start_row": data_start_row,
    }


def _rows_for_sheet(tables: dict[str, dict[str, object]], sheet_name: str) -> list[dict[str, object]]:
    return list(tables.get(sheet_name, {}).get("rows", []))


def _append_check(
    checks: list[dict[str, object]],
    *,
    name: str,
    expected: object,
    actual: object,
    passed: bool,
    details: str,
) -> None:
    checks.append(
        {
            "name": name,
            "expected": expected,
            "actual": actual,
            "passed": passed,
            "details": details,
        }
    )


def _contains_all(actual_headers: object, required_headers: object) -> bool:
    actual = set(actual_headers or [])
    return all(header in actual for header in required_headers or [])


def _row_has_data(values: tuple[object, ...]) -> bool:
    return any(value not in (None, "") for value in values)


def _missing_required_columns(rows: list[dict[str, object]], required_columns: list[str]) -> list[str]:
    missing: set[str] = set()
    for row in rows:
        for column in required_columns:
            if not str(row.get(column) or "").strip():
                missing.add(column)
    return sorted(missing)


def _duplicates(values: list[str]) -> list[str]:
    counts = Counter(value for value in values if value)
    return sorted(value for value, count in counts.items() if count > 1)


def _only_allowed_values(rows: list[dict[str, object]], field: str, allowed_values: list[str]) -> bool:
    allowed = {value.strip() for value in allowed_values}
    actual = {
        str(row.get(field) or "").strip()
        for row in rows
        if str(row.get(field) or "").strip()
    }
    return actual.issubset(allowed)


def _metric_rows_to_map(
    rows: list[dict[str, object]],
    key_field: str,
    value_field: str,
) -> dict[str, object]:
    metrics: dict[str, object] = {}
    for row in rows:
        key = str(row.get(key_field) or "").strip()
        if not key:
            continue
        metrics[key] = row.get(value_field)
    return metrics


def _sum_numeric_field(rows: list[dict[str, object]], field: str) -> int:
    total = 0
    for row in rows:
        total += _coerce_int(row.get(field))
    return total


def _coerce_int(value: object) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(str(value).strip())
