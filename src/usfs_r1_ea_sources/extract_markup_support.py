from __future__ import annotations

from collections import Counter
from html.parser import HTMLParser
from pathlib import Path
import html
import json
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .extract_chunking import _assemble_blocks
from .extract_chunking import _blocks_from_plain_text
from .extract_chunking import _blocks_from_text_fragments
from .extract_common import ExtractionFailure
from .extract_common import ExtractionPayload
from .extract_common import TextBlock
from .extract_common import _clean_text
from .extract_common import _decode_bytes
from .extract_common import _normalize_xml_scope_identifier
from .extract_common import _xml_local_name
from .extract_common import _xml_scope_for_row

def _extract_html(artifact_path: Path) -> ExtractionPayload:
    text = _decode_bytes(artifact_path.read_bytes())
    return _extract_html_text(
        text,
        parser_name="python_htmlparser",
        parser_version=sys.version.split()[0],
    )


def _extract_html_text(text: str, *, parser_name: str, parser_version: str) -> ExtractionPayload:
    parser = _HTMLTextParser()
    try:
        parser.feed(text)
        parser.close()
    except Exception as error:
        raise ExtractionFailure("html_parse_failed", str(error)) from error
    blocks = parser.blocks()
    if not blocks:
        blocks = _blocks_from_text_fragments([_strip_html_text(text)])
    assembled, blocks = _assemble_blocks(blocks)
    return ExtractionPayload(
        text=assembled,
        blocks=blocks,
        parser_name=parser_name,
        parser_version=parser_version,
    )


class _HTMLTextParser(HTMLParser):
    BLOCK_TAGS = {
        "address",
        "article",
        "aside",
        "blockquote",
        "br",
        "caption",
        "dd",
        "div",
        "dt",
        "figcaption",
        "footer",
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "header",
        "li",
        "main",
        "p",
        "pre",
        "section",
        "td",
        "th",
        "title",
        "tr",
    }
    HEADING_TAGS = {"h1", "h2", "h3", "h4", "h5", "h6", "title"}
    SKIP_TAGS = {"script", "style", "noscript", "svg"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._blocks: list[TextBlock] = []
        self._parts: list[str] = []
        self._skip_depth = 0
        self._heading_parts: list[str] | None = None
        self._current_heading: str | None = None

    def handle_starttag(self, tag: str, attrs) -> None:  # noqa: ANN001, ARG002
        tag = tag.lower()
        if tag in self.SKIP_TAGS:
            self._skip_depth += 1
            return
        if tag in self.BLOCK_TAGS:
            self._flush()
        if tag in self.HEADING_TAGS:
            self._heading_parts = []

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in self.SKIP_TAGS and self._skip_depth:
            self._skip_depth -= 1
            return
        if tag in self.HEADING_TAGS:
            heading = _clean_text(" ".join(self._heading_parts or []))
            if heading:
                self._current_heading = heading
            self._flush(heading=heading or self._current_heading)
            self._heading_parts = None
            return
        if tag in self.BLOCK_TAGS:
            self._flush()

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        cleaned = _clean_text(data)
        if not cleaned:
            return
        self._parts.append(cleaned)
        if self._heading_parts is not None:
            self._heading_parts.append(cleaned)

    def blocks(self) -> list[TextBlock]:
        self._flush()
        return self._blocks

    def _flush(self, *, heading: str | None = None) -> None:
        cleaned = _clean_text(" ".join(self._parts))
        self._parts = []
        if not cleaned:
            return
        self._blocks.append(TextBlock(text=cleaned, heading=heading or self._current_heading))


def _strip_html_text(text: str) -> str:
    text = re.sub(r"(?is)<(script|style|noscript|svg).*?</\1>", " ", text)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</(p|div|li|tr|td|th|h[1-6]|section|article)>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    return _clean_text(text)


def _looks_like_html_markup(text: str) -> bool:
    sample = text[:4096].lower()
    return "<html" in sample or "<!doctype html" in sample


def _extract_xml(artifact_path: Path, *, row: dict | None = None) -> ExtractionPayload:
    body = artifact_path.read_bytes()
    try:
        root = ET.fromstring(body)
    except ET.ParseError as error:
        text = _decode_bytes(body)
        if _looks_like_html_markup(text):
            return _extract_html_text(
                text,
                parser_name="legal_xml_xhtml_fallback",
                parser_version=sys.version.split()[0],
            )
        raise ExtractionFailure("xml_parse_failed", str(error)) from error
    source_scope = _xml_scope_for_row(row)
    scoped_root = root
    if source_scope is not None:
        scoped_root = _find_xml_scope_element(root, source_scope)
        if scoped_root is None:
            raise ExtractionFailure(
                "xml_scope_not_found",
                (
                    "Could not find XML element for source scope "
                    f"{source_scope['type']} {source_scope['identifier']}."
                ),
            )
    blocks: list[TextBlock] = []
    _walk_xml(scoped_root, path=(), sibling_index=1, heading=None, blocks=blocks)
    if not blocks:
        blocks = _blocks_from_text_fragments([_clean_text(" ".join(scoped_root.itertext()))])
    assembled, blocks = _assemble_blocks(blocks)
    metadata = {"source_scope": source_scope} if source_scope is not None else None
    return ExtractionPayload(
        text=assembled,
        blocks=blocks,
        parser_name="legal_xml_builtin",
        parser_version=sys.version.split()[0],
        metadata=metadata,
    )


XML_BLOCK_TAGS = {
    "AUTH",
    "CITA",
    "FP",
    "HED",
    "HD",
    "HEAD",
    "NOTE",
    "P",
    "SECTNO",
    "SOURCE",
    "SUBJECT",
}
XML_HEADING_TAGS = {"HED", "HD", "HEAD", "SECTNO", "SUBJECT"}


def _walk_xml(
    element: ET.Element,
    *,
    path: tuple[str, ...],
    sibling_index: int,
    heading: str | None,
    blocks: list[TextBlock],
) -> None:
    tag = _xml_local_name(element.tag)
    element_path = (*path, f"{tag}[{sibling_index}]")
    children = list(element)
    text = _clean_text(" ".join(element.itertext()))
    next_heading = heading
    if text and tag in XML_HEADING_TAGS:
        next_heading = text[:240]
    has_block_child = any(_xml_local_name(child.tag) in XML_BLOCK_TAGS for child in children)
    if text and tag in XML_BLOCK_TAGS and not has_block_child:
        blocks.append(
            TextBlock(
                text=text,
                heading=next_heading,
                section="/" + "/".join(element_path),
            )
        )
    tag_counts: Counter[str] = Counter()
    for child in children:
        child_tag = _xml_local_name(child.tag)
        tag_counts[child_tag] += 1
        _walk_xml(
            child,
            path=element_path,
            sibling_index=tag_counts[child_tag],
            heading=next_heading,
            blocks=blocks,
        )



def _find_xml_scope_element(root: ET.Element, scope: dict) -> ET.Element | None:
    expected_type = _normalize_xml_scope_identifier(scope["type"])
    expected_identifier = _normalize_xml_scope_identifier(scope["identifier"])
    for element in root.iter():
        element_type = _normalize_xml_scope_identifier(element.attrib.get("TYPE", ""))
        element_identifier = _normalize_xml_scope_identifier(element.attrib.get("N", ""))
        if element_type == expected_type and element_identifier == expected_identifier:
            return element
        metadata = _xml_hierarchy_metadata(element)
        if metadata and _metadata_matches_scope(metadata, scope):
            return element
    return None


def _xml_hierarchy_metadata(element: ET.Element) -> dict | None:
    value = element.attrib.get("hierarchy_metadata")
    if not value:
        return None
    try:
        decoded = html.unescape(value)
        metadata = json.loads(decoded)
    except (TypeError, json.JSONDecodeError):
        return None
    return metadata if isinstance(metadata, dict) else None


def _metadata_matches_scope(metadata: dict, scope: dict) -> bool:
    expected_type = str(scope["type"]).lower()
    expected_identifier = str(scope["identifier"]).lower()
    path = str(metadata.get("path") or "").lower()
    citation = str(metadata.get("citation") or "").lower()
    if expected_type == "section":
        return f"/section-{expected_identifier}" in path or citation.endswith(expected_identifier)
    if expected_type == "subpart":
        return f"/subpart-{expected_identifier}" in path or citation.endswith(f"subpart {expected_identifier}")
    return False



def _extract_docx(artifact_path: Path) -> ExtractionPayload:
    try:
        with zipfile.ZipFile(artifact_path) as archive:
            document_xml = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile, OSError) as error:
        raise ExtractionFailure("docx_parse_failed", str(error)) from error

    try:
        root = ET.fromstring(document_xml)
    except ET.ParseError as error:
        raise ExtractionFailure("docx_parse_failed", str(error)) from error

    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraph_tag = f"{namespace}p"
    text_tag = f"{namespace}t"
    fragments: list[str] = []
    for paragraph in root.iter(paragraph_tag):
        paragraph_text = "".join(text_node.text or "" for text_node in paragraph.iter(text_tag))
        cleaned = _clean_text(paragraph_text)
        if cleaned:
            fragments.append(cleaned)
    text, blocks = _blocks_from_plain_text("\n\n".join(fragments))
    return ExtractionPayload(
        text=text,
        blocks=blocks,
        parser_name="python_docx_zip_xml",
        parser_version=sys.version.split()[0],
    )


def _extract_xlsx(artifact_path: Path) -> ExtractionPayload:
    try:
        workbook = load_workbook(artifact_path, read_only=True, data_only=True)
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        raise ExtractionFailure("xlsx_parse_failed", str(error)) from error

    blocks: list[TextBlock] = []
    rendered_rows: list[str] = []
    sheet_summaries: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        nonempty_row_count = 0
        for row_index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            rendered_cells = [_xlsx_cell_text(value) for value in values]
            if not any(rendered_cells):
                continue
            nonempty_row_count += 1
            last_column_index = max(
                index for index, value in enumerate(rendered_cells, start=1) if value
            )
            text = " | ".join(value for value in rendered_cells if value)
            section = f"{sheet.title}!A{row_index}:{get_column_letter(last_column_index)}{row_index}"
            blocks.append(TextBlock(text=text, heading=sheet.title, section=section))
            rendered_rows.append(f"[{section}] {text}")
        sheet_summaries.append(
            {
                "sheet_name": sheet.title,
                "nonempty_row_count": nonempty_row_count,
            }
        )

    if not blocks:
        raise ExtractionFailure(
            "xlsx_parse_failed",
            "Workbook did not contain any non-empty readable cell rows.",
        )

    return ExtractionPayload(
        text="\n".join(rendered_rows),
        blocks=blocks,
        parser_name="openpyxl_xlsx_cells",
        parser_version=sys.version.split()[0],
        metadata={
            "sheet_count": len(workbook.worksheets),
            "sheets": sheet_summaries,
        },
    )


def _extract_zip(artifact_path: Path) -> ExtractionPayload:
    try:
        with zipfile.ZipFile(artifact_path) as archive:
            entry_names = sorted(archive.namelist())
            metadata_name = next(
                (name for name in entry_names if name.lower().endswith(".xml")),
                None,
            )
            metadata_bytes = archive.read(metadata_name) if metadata_name else None
    except (KeyError, zipfile.BadZipFile, OSError) as error:
        raise ExtractionFailure("zip_parse_failed", str(error)) from error

    fragments = ["Archive entries:\n" + "\n".join(entry_names)]
    if metadata_bytes:
        metadata_text = _extract_zip_metadata_text(metadata_bytes)
        if metadata_text:
            fragments.append(metadata_text)
    text, blocks = _blocks_from_plain_text("\n\n".join(fragment for fragment in fragments if fragment))
    return ExtractionPayload(
        text=text,
        blocks=blocks,
        parser_name="python_zip_metadata_xml",
        parser_version=sys.version.split()[0],
        metadata={
            "archive_entry_count": len(entry_names),
            "metadata_entry": metadata_name,
        },
    )


def _extract_zip_metadata_text(payload: bytes) -> str:
    decoded = payload.decode("utf-8", errors="ignore")
    try:
        root = ET.fromstring(decoded)
    except ET.ParseError:
        return _clean_text(re.sub(r"<[^>]+>", " ", decoded))

    fragments: list[str] = []
    seen: set[str] = set()
    for element in root.iter():
        text = _clean_text(element.text or "")
        if not text:
            continue
        tag = _xml_local_name(element.tag)
        if tag.lower() in {"wkt", "pexml"}:
            continue
        if len(text) > 500:
            text = text[:500].rstrip() + "..."
        fragment = f"{tag}: {text}" if tag else text
        if fragment in seen:
            continue
        seen.add(fragment)
        fragments.append(fragment)
    return "\n".join(fragments)


def _extract_doc(artifact_path: Path, *, facade_module: Any) -> ExtractionPayload:
    command = ["textutil", "-convert", "txt", "-stdout", str(artifact_path)]
    try:
        completed = facade_module.subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
        )
    except OSError as error:
        raise ExtractionFailure("doc_parse_failed", str(error)) from error
    if completed.returncode != 0:
        raise ExtractionFailure(
            "doc_parse_failed",
            (completed.stderr or completed.stdout or "textutil failed").strip(),
        )
    text = completed.stdout.strip()
    if not text:
        raise ExtractionFailure("doc_parse_failed", "textutil returned empty text.")
    text, blocks = _blocks_from_plain_text(text)
    return ExtractionPayload(
        text=text,
        blocks=blocks,
        parser_name="macos_textutil_doc",
        parser_version="system",
    )


def _extract_image(
    artifact_path: Path,
    *,
    ocr_enabled: bool,
    timeout_seconds: float | None,
    facade_module: Any,
) -> ExtractionPayload:
    payload = facade_module._try_extract_docling(
        artifact_path,
        ocr_enabled=True if not ocr_enabled else ocr_enabled,
        timeout_seconds=timeout_seconds,
    )
    if payload is None:
        raise ExtractionFailure(
            "docling_unavailable",
            "Image extraction requires Docling; set USFS_R1_DOCLING_PYTHON or run from a Docling environment.",
        )
    return payload


def _extract_plain_text(artifact_path: Path) -> ExtractionPayload:
    text, blocks = _blocks_from_plain_text(_decode_bytes(artifact_path.read_bytes()))
    return ExtractionPayload(
        text=text,
        blocks=blocks,
        parser_name="python_text_decode",
        parser_version=sys.version.split()[0],
    )


def _xlsx_cell_text(value: object) -> str:
    if value in (None, ""):
        return ""
    return " ".join(str(value).strip().split())
