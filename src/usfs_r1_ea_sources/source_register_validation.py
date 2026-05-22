from __future__ import annotations

from importlib import import_module
from pathlib import Path

from openpyxl import load_workbook


def validate_source_register(
    workbook_path: Path | str,
    *,
    mode: str = "schema",
    sheet_contract_path=None,
    schema_path=None,
    vocabularies_path=None,
    row_states_path=None,
) -> dict:
    source_register_module = import_module("usfs_r1_ea_sources.source_register")

    if sheet_contract_path is None:
        sheet_contract_path = source_register_module.DEFAULT_SOURCE_REGISTER_SHEET_CONTRACT_PATH
    if schema_path is None:
        schema_path = source_register_module.DEFAULT_SOURCE_REGISTER_SCHEMA_PATH
    if vocabularies_path is None:
        vocabularies_path = source_register_module.DEFAULT_SOURCE_REGISTER_VOCABULARIES_PATH
    if row_states_path is None:
        row_states_path = source_register_module.DEFAULT_SOURCE_REGISTER_ROW_STATES_PATH

    if mode != "schema":
        raise ValueError(f"Unsupported source register validation mode: {mode!r}")

    workbook_path = Path(workbook_path)
    contract = source_register_module.load_source_register_contract(sheet_contract_path)
    schema = source_register_module.load_source_register_schema(schema_path)
    vocabularies = source_register_module.load_source_register_vocabularies(vocabularies_path)
    row_states = source_register_module.load_source_register_row_states(row_states_path)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    tables = source_register_module.read_source_register_tables(
        workbook_path,
        contract_path=sheet_contract_path,
    )
    checks: list[dict[str, object]] = []

    _append_check(
        checks,
        name="sheet_roster_matches_contract",
        expected=contract["sheet_order"],
        actual=workbook.sheetnames,
        passed=workbook.sheetnames == contract["sheet_order"],
        details="Workbook sheet order must match the frozen Phase 0 sheet contract.",
    )

    for spec in contract["sheets"]:
        sheet_name = spec["sheet_name"]
        table = tables.get(sheet_name)
        if table is None:
            _append_check(
                checks,
                name=f"{sheet_name}_present",
                expected=True,
                actual=False,
                passed=False,
                details="Required sheet is missing from the canonical source register.",
            )
            continue
        _append_check(
            checks,
            name=f"{sheet_name}_required_columns_present",
            expected=spec["required_columns"],
            actual=table["headers"],
            passed=_contains_all(table["headers"], spec["required_columns"]),
            details="Required columns must exist at the frozen header row.",
        )

    master_rows = source_register_module._rows_for_sheet(tables, contract["load_sheet_name"])
    queue_rows = source_register_module._rows_for_sheet(tables, contract["queue_sheet_name"])
    removed_rows = source_register_module._rows_for_sheet(tables, contract["removed_sheet_name"])

    master_schema = schema["master_sheet"]
    queue_schema = schema["queue_sheet"]
    removed_schema = schema["removed_sheet"]
    vocabulary_values = vocabularies["controlled_values"]

    master_source_ids = [str(row["Source_ID"]).strip() for row in master_rows]
    master_urls = [str(row["Source_URL"]).strip() for row in master_rows]
    master_normalized_urls = [source_register_module.normalize_url(url) for url in master_urls if url]
    blank_master_urls = [row["Source_ID"] for row in master_rows if not str(row["Source_URL"] or "").strip()]
    non_http_master_urls = [
        row["Source_ID"]
        for row in master_rows
        if source_register_module.urlsplit(str(row["Source_URL"] or "").strip()).scheme not in {"http", "https"}
    ]
    missing_master_applicability = [
        row["Source_ID"]
        for row in master_rows
        if not str(row.get("EA_System_Applicability_Status") or "").strip()
    ]
    duplicate_source_ids = _duplicates(master_source_ids)
    duplicate_source_urls = _duplicates(master_normalized_urls)

    _append_check(
        checks,
        name="master_required_nonempty_columns",
        expected=master_schema["required_nonempty_columns"],
        actual=_missing_required_columns(master_rows, master_schema["required_nonempty_columns"]),
        passed=not _missing_required_columns(master_rows, master_schema["required_nonempty_columns"]),
        details="Required master-sheet fields must be populated for every retained load row.",
    )
    _append_check(
        checks,
        name="master_source_id_unique",
        expected=0,
        actual=len(duplicate_source_ids),
        passed=not duplicate_source_ids,
        details="Document_Register_Master Source_ID values must be unique.",
    )
    _append_check(
        checks,
        name="master_source_url_unique",
        expected=0,
        actual=len(duplicate_source_urls),
        passed=not duplicate_source_urls,
        details="Document_Register_Master Source_URL values must be unique after normalization.",
    )
    _append_check(
        checks,
        name="master_source_url_nonempty",
        expected=0,
        actual=len(blank_master_urls),
        passed=not blank_master_urls,
        details="Document_Register_Master rows must not have blank Source_URL values.",
    )
    _append_check(
        checks,
        name="master_source_url_http",
        expected=0,
        actual=len(non_http_master_urls),
        passed=not non_http_master_urls,
        details="Document_Register_Master rows must use HTTP(S) source URLs.",
    )
    _append_check(
        checks,
        name="master_applicability_complete",
        expected=0,
        actual=len(missing_master_applicability),
        passed=not missing_master_applicability,
        details="Document_Register_Master rows must carry EA_System_Applicability_Status.",
    )
    _append_check(
        checks,
        name="master_database_load_values",
        expected=vocabulary_values["master_database_load"],
        actual=sorted({str(row["Database_Load"]).strip() for row in master_rows}),
        passed=_only_allowed_values(
            master_rows,
            "Database_Load",
            vocabulary_values["master_database_load"],
        ),
        details="Retained master rows must stay in the load-ready state.",
    )
    _append_check(
        checks,
        name="master_authority_tier_values",
        expected=vocabulary_values["master_authority_tier"],
        actual=sorted({str(row["Authority_Tier"]).strip() for row in master_rows}),
        passed=_only_allowed_values(
            master_rows,
            "Authority_Tier",
            vocabulary_values["master_authority_tier"],
        ),
        details="Master-sheet authority tiers must stay within the frozen vocabulary.",
    )
    _append_check(
        checks,
        name="master_applicability_status_values",
        expected=vocabulary_values["load_applicability_statuses"],
        actual=sorted({str(row["EA_System_Applicability_Status"]).strip() for row in master_rows}),
        passed=_only_allowed_values(
            master_rows,
            "EA_System_Applicability_Status",
            vocabulary_values["load_applicability_statuses"],
        ),
        details="Master-sheet applicability statuses must stay within the frozen vocabulary.",
    )

    _append_check(
        checks,
        name="queue_required_nonempty_columns",
        expected=queue_schema["required_nonempty_columns"],
        actual=_missing_required_columns(queue_rows, queue_schema["required_nonempty_columns"]),
        passed=not _missing_required_columns(queue_rows, queue_schema["required_nonempty_columns"]),
        details="Queue rows must preserve the fields needed for later direct-file promotion.",
    )
    _append_check(
        checks,
        name="queue_database_load_values",
        expected=vocabulary_values["queue_database_load"],
        actual=sorted({str(row["Database_Load"]).strip() for row in queue_rows}),
        passed=_only_allowed_values(
            queue_rows,
            "Database_Load",
            vocabulary_values["queue_database_load"],
        ),
        details="Direct-file capture queue rows must remain non-load rows.",
    )
    _append_check(
        checks,
        name="queue_authority_tier_values",
        expected=vocabulary_values["queue_authority_tier"],
        actual=sorted({str(row["Authority_Tier"]).strip() for row in queue_rows}),
        passed=_only_allowed_values(
            queue_rows,
            "Authority_Tier",
            vocabulary_values["queue_authority_tier"],
        ),
        details="Queue-sheet authority tiers must stay within the frozen vocabulary.",
    )

    _append_check(
        checks,
        name="removed_required_nonempty_columns",
        expected=removed_schema["required_nonempty_columns"],
        actual=_missing_required_columns(removed_rows, removed_schema["required_nonempty_columns"]),
        passed=not _missing_required_columns(removed_rows, removed_schema["required_nonempty_columns"]),
        details="Removed rows must retain removal rationale and applicability state.",
    )
    _append_check(
        checks,
        name="removed_applicability_status_values",
        expected=vocabulary_values["removed_applicability_statuses"],
        actual=sorted({str(row["EA_System_Applicability_Status"]).strip() for row in removed_rows}),
        passed=_only_allowed_values(
            removed_rows,
            "EA_System_Applicability_Status",
            vocabulary_values["removed_applicability_statuses"],
        ),
        details="Removed rows must remain explicitly non-load audit rows.",
    )

    codex_metrics = _metric_rows_to_map(
        source_register_module._rows_for_sheet(tables, "Codex_Load_Index"),
        "Metric",
        "Value",
    )
    final_ingest_metrics = _metric_rows_to_map(
        source_register_module._rows_for_sheet(tables, "Final_Ingest_Certification_2026"),
        "Metric",
        "Value",
    )
    link_summary = _metric_rows_to_map(
        source_register_module._rows_for_sheet(tables, "Link_Validation_Summary"),
        "Category",
        "Count",
    )
    authority_tier_total = _sum_numeric_field(
        source_register_module._rows_for_sheet(tables, "Authority_Tier_Index"),
        "Database_Load_Rows",
    )
    applicability_total = _sum_numeric_field(
        source_register_module._rows_for_sheet(tables, "Applicability_By_Tier"),
        "Row_Count",
    )

    _append_check(
        checks,
        name="codex_master_count_matches_actual",
        expected=len(master_rows),
        actual=_coerce_int(codex_metrics.get("Master database source rows")),
        passed=_coerce_int(codex_metrics.get("Master database source rows")) == len(master_rows),
        details="Codex_Load_Index must match the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="codex_queue_count_matches_actual",
        expected=len(queue_rows),
        actual=_coerce_int(codex_metrics.get("Capture queue rows")),
        passed=_coerce_int(codex_metrics.get("Capture queue rows")) == len(queue_rows),
        details="Codex_Load_Index must match the actual queue-row count.",
    )
    _append_check(
        checks,
        name="codex_removed_count_matches_actual",
        expected=len(removed_rows),
        actual=_coerce_int(codex_metrics.get("Rows removed in this pass")),
        passed=_coerce_int(codex_metrics.get("Rows removed in this pass")) == len(removed_rows),
        details="Codex_Load_Index must match the actual removed-row count.",
    )
    _append_check(
        checks,
        name="final_ingest_master_count_matches_actual",
        expected=len(master_rows),
        actual=_coerce_int(final_ingest_metrics.get("Final retained database-load rows")),
        passed=_coerce_int(final_ingest_metrics.get("Final retained database-load rows")) == len(master_rows),
        details="Final_Ingest_Certification_2026 must match the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="link_summary_master_count_matches_actual",
        expected=len(master_rows),
        actual=_coerce_int(link_summary.get("Working official links in master")),
        passed=_coerce_int(link_summary.get("Working official links in master")) == len(master_rows),
        details="Link_Validation_Summary must match the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="link_summary_queue_count_matches_actual",
        expected=len(queue_rows),
        actual=_coerce_int(link_summary.get("Capture queue non-load rows")),
        passed=_coerce_int(link_summary.get("Capture queue non-load rows")) == len(queue_rows),
        details="Link_Validation_Summary must match the actual queue-row count.",
    )
    _append_check(
        checks,
        name="authority_tier_index_total_matches_actual",
        expected=len(master_rows),
        actual=authority_tier_total,
        passed=authority_tier_total == len(master_rows),
        details="Authority_Tier_Index totals must reconcile to the actual retained master-row count.",
    )
    _append_check(
        checks,
        name="applicability_by_tier_total_matches_actual",
        expected=len(master_rows),
        actual=applicability_total,
        passed=applicability_total == len(master_rows),
        details="Applicability_By_Tier totals must reconcile to the actual retained master-row count.",
    )

    validation_passed = all(bool(check["passed"]) for check in checks)
    state_sheet_names = {
        state["state_id"]: state.get("sheet_name") or state.get("sheet_names")
        for state in row_states["states"]
    }
    return {
        "schema_version": "source-register-validation-v1",
        "mode": mode,
        "workbook_path": str(workbook_path),
        "workbook_sha256": source_register_module.sha256_file(workbook_path),
        "sheet_contract_path": str(Path(sheet_contract_path)),
        "schema_path": str(Path(schema_path)),
        "vocabularies_path": str(Path(vocabularies_path)),
        "row_states_path": str(Path(row_states_path)),
        "sheet_count": len(workbook.sheetnames),
        "load_sheet_name": contract["load_sheet_name"],
        "queue_sheet_name": contract["queue_sheet_name"],
        "removed_sheet_name": contract["removed_sheet_name"],
        "load_row_count": len(master_rows),
        "queue_row_count": len(queue_rows),
        "removed_row_count": len(removed_rows),
        "stale_source_detector_count": sum(
            1
            for row in master_rows
            if str(row.get("EA_System_Applicability_Status") or "").strip()
            == "Applicable - stale-source detector only"
        ),
        "row_state_sheet_names": state_sheet_names,
        "checks": checks,
        "issue_count": sum(1 for check in checks if not check["passed"]),
        "validation_passed": validation_passed,
    }


def _append_check(
    checks: list[dict[str, object]],
    *,
    name: str,
    expected: object,
    actual: object,
    passed: bool,
    details: str,
) -> None:
    checks.append(
        {
            "name": name,
            "expected": expected,
            "actual": actual,
            "passed": passed,
            "details": details,
        }
    )


def _contains_all(actual_headers: object, required_headers: object) -> bool:
    actual = set(actual_headers or [])
    return all(header in actual for header in required_headers or [])


def _missing_required_columns(rows: list[dict[str, object]], required_columns: list[str]) -> list[str]:
    missing: set[str] = set()
    for row in rows:
        for column in required_columns:
            if not str(row.get(column) or "").strip():
                missing.add(column)
    return sorted(missing)


def _duplicates(values: list[str]) -> list[str]:
    from collections import Counter

    counts = Counter(value for value in values if value)
    return sorted(value for value, count in counts.items() if count > 1)


def _only_allowed_values(rows: list[dict[str, object]], field: str, allowed_values: list[str]) -> bool:
    allowed = {value.strip() for value in allowed_values}
    actual = {
        str(row.get(field) or "").strip()
        for row in rows
        if str(row.get(field) or "").strip()
    }
    return actual.issubset(allowed)


def _metric_rows_to_map(
    rows: list[dict[str, object]],
    key_field: str,
    value_field: str,
) -> dict[str, object]:
    metrics: dict[str, object] = {}
    for row in rows:
        key = str(row.get(key_field) or "").strip()
        if not key:
            continue
        metrics[key] = row.get(value_field)
    return metrics


def _sum_numeric_field(rows: list[dict[str, object]], field: str) -> int:
    total = 0
    for row in rows:
        total += _coerce_int(row.get(field))
    return total


def _coerce_int(value: object) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(str(value).strip())
