from __future__ import annotations

from pathlib import Path
import shutil
import tempfile
import unittest

from openpyxl import load_workbook

from usfs_r1_ea_sources.source_register import build_source_register_diff
from usfs_r1_ea_sources.source_register import validate_source_register


ROOT = Path(__file__).resolve().parents[1]
CANONICAL_WORKBOOK = ROOT / "usfs_region1_ea_source_register_FINAL_INGEST_READY_2026.xlsx"
LEGACY_WORKBOOK = ROOT / "usfs_region1_ea_document_checklist_land_exchange_review_2026.xlsx"
LEGACY_REGISTER = ROOT / "config" / "r1_forest_plan_document_register_draft.csv"


class SourceRegisterSchemaTests(unittest.TestCase):
    def test_validate_source_register_passes_for_final_workbook(self) -> None:
        result = validate_source_register(CANONICAL_WORKBOOK)

        self.assertTrue(result["validation_passed"])
        self.assertEqual(result["sheet_count"], 13)
        self.assertEqual(result["load_row_count"], 691)
        self.assertEqual(result["queue_row_count"], 51)
        self.assertEqual(result["removed_row_count"], 3)
        self.assertEqual(result["stale_source_detector_count"], 5)

    def test_validate_source_register_detects_duplicate_load_url(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook_copy = Path(tmp) / CANONICAL_WORKBOOK.name
            shutil.copyfile(CANONICAL_WORKBOOK, workbook_copy)
            workbook = load_workbook(workbook_copy)
            worksheet = workbook["Document_Register_Master"]
            headers = _header_map(worksheet)
            source_url_column = headers["Source_URL"]
            worksheet.cell(row=6, column=source_url_column).value = worksheet.cell(
                row=5,
                column=source_url_column,
            ).value
            workbook.save(workbook_copy)

            result = validate_source_register(workbook_copy)

        failing_checks = {check["name"] for check in result["checks"] if not check["passed"]}
        self.assertFalse(result["validation_passed"])
        self.assertIn("master_source_url_unique", failing_checks)

    def test_validate_source_register_detects_queue_database_load_leakage(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook_copy = Path(tmp) / CANONICAL_WORKBOOK.name
            shutil.copyfile(CANONICAL_WORKBOOK, workbook_copy)
            workbook = load_workbook(workbook_copy)
            worksheet = workbook["Direct_File_Capture_Queue"]
            headers = _header_map(worksheet)
            worksheet.cell(row=5, column=headers["Database_Load"]).value = "Yes"
            workbook.save(workbook_copy)

            result = validate_source_register(workbook_copy)

        failing_checks = {check["name"] for check in result["checks"] if not check["passed"]}
        self.assertFalse(result["validation_passed"])
        self.assertIn("queue_database_load_values", failing_checks)

    def test_build_source_register_diff_reports_phase_zero_replacement_counts(self) -> None:
        result = build_source_register_diff(
            LEGACY_WORKBOOK,
            LEGACY_REGISTER,
            CANONICAL_WORKBOOK,
        )

        self.assertEqual(result["legacy_workbook_source_count"], 190)
        self.assertEqual(result["legacy_register_source_delta_count"], 160)
        self.assertEqual(result["legacy_register_gap_count"], 1)
        self.assertEqual(result["legacy_runtime_unique_source_count"], 350)
        self.assertEqual(result["canonical_master_row_count"], 691)
        self.assertEqual(result["canonical_queue_row_count"], 51)
        self.assertEqual(result["canonical_removed_row_count"], 3)
        self.assertEqual(result["canonical_stale_source_detector_count"], 5)
        self.assertEqual(result["canonical_shared_with_legacy_workbook_count"], 12)
        self.assertEqual(result["canonical_shared_with_source_delta_count"], 0)
        self.assertEqual(result["canonical_only_source_count"], 679)
        self.assertEqual(result["legacy_only_source_count"], 338)
        self.assertEqual(result["canonical_only_source_ids_sample"][0], "FED-001")
        self.assertEqual(result["legacy_only_source_ids_sample"][0], "R1EA-001")

    def test_fps_005_is_removed_from_active_ingest_with_governed_reason(self) -> None:
        workbook = load_workbook(CANONICAL_WORKBOOK, read_only=True, data_only=True)
        master = workbook["Document_Register_Master"]
        removed = workbook["Removed_Not_Applicable_Final"]
        master_headers = _header_map(master)
        removed_headers = _header_map(removed)

        master_ids = {
            str(row[master_headers["Source_ID"] - 1])
            for row in master.iter_rows(min_row=5, values_only=True)
            if row[master_headers["Source_ID"] - 1]
        }
        self.assertNotIn("FPS-005", master_ids)

        removed_rows = {
            str(row[removed_headers["Source_ID"] - 1]): row
            for row in removed.iter_rows(min_row=5, values_only=True)
            if row[removed_headers["Source_ID"] - 1]
        }
        fps_005 = removed_rows["FPS-005"]
        self.assertEqual(
            fps_005[removed_headers["Source_URL"] - 1],
            "https://www.fs.usda.gov/media/228272",
        )
        self.assertEqual(
            fps_005[removed_headers["EA_System_Applicability_Status"] - 1],
            "Not applicable - removed from ingest",
        )
        self.assertIn(
            "structurally invalid",
            str(fps_005[removed_headers["Removal_Reason"] - 1]),
        )

    def test_known_blocker_repairs_use_current_official_urls(self) -> None:
        workbook = load_workbook(CANONICAL_WORKBOOK, read_only=True, data_only=True)
        worksheet = workbook["Document_Register_Master"]
        headers = _header_map(worksheet)
        source_id_column = headers["Source_ID"]
        source_url_column = headers["Source_URL"]
        expected_urls = {
            "FED-042": "https://www.archives.gov/federal-register/codification/executive-order/11988.html",
            "FED-041": "https://www.govinfo.gov/content/pkg/FR-1999-02-08/html/99-3184.htm",
            "FED-039": "https://www.govinfo.gov/content/pkg/FR-1996-05-29/html/96-13597.htm",
            "FED-043": "https://www.archives.gov/federal-register/codification/executive-order/11990.html",
            "FED-044": "https://www.ecfr.gov/current/title-40/chapter-I/subchapter-C/part-93/subpart-B",
            "FED-045": "https://uscode.house.gov/view.xhtml?req=(title:33%20section:1341%20edition:prelim)",
            "FED-046": "https://uscode.house.gov/view.xhtml?req=(title:33%20section:1342%20edition:prelim)",
            "FED-047": "https://uscode.house.gov/view.xhtml?req=(title:33%20section:1344%20edition:prelim)",
            "FED-048": "https://uscode.house.gov/view.xhtml?req=(title:33%20section:403%20edition:prelim)",
            "FED-049": "https://www.ecfr.gov/current/title-33/chapter-II/part-320",
            "FED-050": "https://www.ecfr.gov/current/title-33/chapter-II/part-322",
            "FED-051": "https://www.ecfr.gov/current/title-33/chapter-II/part-325",
            "FED-052": "https://uscode.house.gov/view.xhtml?req=granuleid%3AUSC-prelim-title54-section306102&num=0&edition=prelim",
            "FED-053": "https://www.ecfr.gov/current/title-36/chapter-II/part-296",
            "FED-054": "https://www.ecfr.gov/current/title-43/subtitle-A/part-10",
            "FED-055": "https://uscode.house.gov/view.xhtml?path=/prelim@title25/chapter32A&edition=prelim",
            "FED-056": "https://uscode.house.gov/view.xhtml?req=(title:42%20section:1996%20edition:prelim)",
            "FED-057": "https://uscode.house.gov/view.xhtml?path=/prelim@title42/chapter21B&edition=prelim",
            "FED-058": "https://uscode.house.gov/view.xhtml?req=(title:16%20section:470aaa%20edition:prelim)",
            "FED-059": "https://uscode.house.gov/view.xhtml?path=/prelim@title16/chapter63&edition=prelim",
            "FED-029": "https://uscode.house.gov/view.xhtml?path=/prelim@title16/chapter5A&edition=prelim",
            "FED-060": "https://www.ecfr.gov/current/title-50/chapter-I/subchapter-B/part-22",
            "FED-061": "https://uscode.house.gov/view.xhtml?req=(title:16%20section:1855%20edition:prelim)",
            "FED-062": "https://www.ecfr.gov/current/title-50/chapter-VI/part-600/subpart-K",
            "FED-063": "https://www.ecfr.gov/current/title-40/chapter-I/subchapter-J/part-300",
            "FED-064": "https://uscode.house.gov/view.xhtml?path=/prelim@title7/chapter73&edition=prelim",
            "FED-065": "https://www.ecfr.gov/current/title-7/subtitle-B/chapter-VI/subchapter-F/part-658",
            "FED-066": "https://uscode.house.gov/view.xhtml?path=/prelim@title42/chapter6A/subchapter12&edition=prelim",
            "FED-067": "https://www.ecfr.gov/current/title-40/chapter-I/subchapter-D/part-144",
            "FED-068": "https://www.epa.gov/enforcement/federal-insecticide-fungicide-and-rodenticide-act-fifra-and-federal-facilities",
            "FED-069": "https://www.federalregister.gov/documents/2016/12/08/2016-29519/safeguarding-the-nation-from-the-impacts-of-invasive-species",
            "FED-070": "https://www.federalregister.gov/documents/2025/03/25/2025-05212/immediate-measures-to-increase-american-mineral-production",
            "R1PLAN-region-1-northern-region-02": "https://www.fs.usda.gov/r01/forests-grasslands",
            "R1PLAN-beaverhead-deerlodge-nf-01": "https://www.fs.usda.gov/r01/beaverhead-deerlodge/planning",
            "R1PLAN-bitterroot-nf-01": "https://www.fs.usda.gov/r01/bitterroot/planning",
            "R1PLAN-custer-gallatin-nf-01": "https://www.fs.usda.gov/r01/custergallatin/planning/forest-plan/custer-gallatin-land-management-plan-forest-plan-revision",
            "R1PLAN-dakota-prairie-grasslands-01": "https://www.fs.usda.gov/r01/dpg/natural-resources",
            "R1PLAN-flathead-nf-01": "https://www.fs.usda.gov/r01/flathead/planning/forest-plan",
            "R1PLAN-helena-lewis-and-clark-nf-01": "https://www.fs.usda.gov/r01/helena-lewisclark/natural-resources/forest-management/2021-forest-plan-final-eis-and-record-0",
            "R1PLAN-idaho-panhandle-nfs-01": "https://www.fs.usda.gov/r01/idahopanhandle/planning",
            "R1PLAN-kootenai-nf-01": "https://www.fs.usda.gov/r01/kootenai/planning",
            "R1PLAN-lolo-nf-01": "https://www.fs.usda.gov/r01/lolo/planning",
            "R1PLAN-nez-perce-clearwater-nfs-01": "https://www.fs.usda.gov/r01/nezperce-clearwater/planning",
            "R1PLAN-nez-perce-clearwater-nfs-02": "https://www.fs.usda.gov/r01/nezperce-clearwater/planning/2025-land-management-plan",
            "FPS-117": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/custergallatin/CNF%20FPAdjustment%20001.pdf",
            "FINAL-Q-HLC-001": "https://www.fs.usda.gov/sites/nfs/files/r01/helena-lewisclark/publication/V3%20Maps%20EIS%202021%20Forest%20Plan.pdf",
            "FINAL-Q-HLC-002": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/helena-lewisclark/Volume%204%20-%20HLCNF%20Plan.pdf",
            "FINAL-Q-HLC-003": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/helena-lewisclark/Volume%205%20-%20HLCNF%20Plan.pdf",
            "FPS-344": "https://www.federalregister.gov/d/2024-30342",
            "PROG-008": "https://www.fs.usda.gov/naturalresources/watershed/pubs/FS_National_Core_BMPs_April2012.pdf",
            "PROG-010": "https://fs-prod-nwcg.s3.us-gov-west-1.amazonaws.com/s3fs-public/publication/pms484.pdf",
            "R1-SCC-CGNF-005": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/20210415_CG%20SCC%20Animal%20Rational_Objection%20Response.xlsx",
            "R1-SCC-CGNF-006": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/20210415_CG%20SCC%20Plant%20Rational_Objection%20Response.xlsx",
            "R1-SCC-FLAT-005": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/2018-11-09_SCC%20Terrestrial%20Wildlife%20Eval%20Post%20Objection_Flathead.xlsx",
            "R1-SCC-FLAT-006": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/2018-11-09_SCC%20Aquatic%20Animal%20Post%20Objection_Flathead.xlsx",
            "R1-SCC-FLAT-007": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/2018-11-09_SCC%20Plant%20Evaluations%20Revision_Flathead.xlsx",
            "R1-SCC-HLC-005": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/2021-02-19%20HLC%20SCC_Animal%20Rational_Final_Post%20Objection.xlsx",
            "R1-SCC-HLC-006": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/2021-02-19%20HLC%20SCC_Plant%20Rational_Final_Post%20Objection.xlsx",
            "R1-SCC-NPC-004": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/NPC%20FPR%20SCC%20Animals%20January%202025.xlsx",
            "R1-SCC-NPC-005": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/NPC%20FPR%20SCC%20Plants%20November%202023.xlsx",
            "STP-015": "https://www.deq.idaho.gov/water-quality/surface-water/total-maximum-daily-loads/",
            "STP-031": "https://deq.mt.gov/water/assistance",
            "STP-032": "https://www.deq.idaho.gov/permits/water-quality-permits-certifications/water-quality-certifications/",
            "STP-033": "https://deq.nd.gov/WQ/5_Special_Projects/default.aspx",
            "STP-034": "https://danr.sd.gov/OfficeOfWater/SurfaceWaterQuality/waterqualitystandards/401Certifications.aspx",
            "STP-035": "https://dahp.wa.gov/project-review",
            "STP-011": "https://idfg.idaho.gov/species/",
            "USDA-008": "https://www.fs.usda.gov/im/directives/",
            "USDA-009": "https://www.ams.usda.gov/about-ams/accessibility",
            "USDA-010": "https://securefoia.usda.gov/",
            "USDA-011": "https://www.fs.usda.gov/spf/tribalrelations/documents/policy/consultation/Final_DR.pdf",
            "USDA-012": "https://www.rma.usda.gov/about-rma/website-policies-important-links/nondiscrimination-statement",
            "USDA-013": "https://www.ers.usda.gov/about-ers/policies-and-standards/information-quality",
            "WILD-ESA-094": "https://www.fs.usda.gov/sites/nfs/files/legacy-media/r01/lynx%20mgmt%20dir%20veg%20small%20map.jpg",
        }
        actual_urls: dict[str, str] = {}

        for row in worksheet.iter_rows(min_row=5, values_only=True):
            source_id = row[source_id_column - 1]
            if source_id in expected_urls:
                actual_urls[str(source_id)] = str(row[source_url_column - 1])

        self.assertEqual(actual_urls, expected_urls)


def _header_map(worksheet) -> dict[str, int]:  # noqa: ANN001
    return {
        str(cell.value): index
        for index, cell in enumerate(worksheet[4], start=1)
        if cell.value not in (None, "")
    }


if __name__ == "__main__":
    unittest.main()
