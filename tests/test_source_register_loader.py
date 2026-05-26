from __future__ import annotations

from dataclasses import replace
from pathlib import Path
import tempfile

from openpyxl import load_workbook

from usfs_r1_ea_sources.config import (
    SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT,
    load_config,
)
from usfs_r1_ea_sources.source_register import load_source_register_rows
from usfs_r1_ea_sources.workbook import load_canonical_sources


ROOT = Path(__file__).resolve().parents[1]
CONFIG = ROOT / "config" / "downloader.toml"
CANONICAL_WORKBOOK = ROOT / "usfs_region1_ea_source_register_FINAL_INGEST_READY_2026.xlsx"


def test_active_config_uses_canonical_source_register_loader_contract() -> None:
    config = load_config(CONFIG)
    assert config.workbook.loader_contract == SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT


def test_source_register_loader_dispatch_returns_workbook_source_compatibility_rows() -> None:
    config = load_config(CONFIG)
    workbook_config = replace(
        config.workbook,
        loader_contract=SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT,
        overrides_path=None,
    )

    sources = load_canonical_sources(CANONICAL_WORKBOOK, workbook_config)

    assert len(sources) == 708
    assert all(source.sheet == "Document_Register_Master" for source in sources)
    assert all(source.metadata["loader_contract"] == SOURCE_REGISTER_WORKBOOK_LOADER_CONTRACT for source in sources)
    assert all(source.metadata["row_state"] == "load_ready_master_row" for source in sources)
    assert all(source.metadata["direct_file_readiness_class"] == "load_ready" for source in sources)
    assert all(source.metadata["authority_document_id"] for source in sources)
    assert all(source.metadata["source_authority_link_id"] for source in sources)


def test_active_canonical_loader_ignores_legacy_override_registry() -> None:
    config = load_config(CONFIG)
    sources = load_canonical_sources(CANONICAL_WORKBOOK, config.workbook)

    nepa_source = next(source for source in sources if source.source_record_id == "FED-001")
    assert nepa_source.effective_url == nepa_source.original_url
    assert "override_url" not in nepa_source.metadata
    assert "override_reason" not in nepa_source.metadata


def test_source_register_loader_exposes_semantic_identity_and_scope_seams() -> None:
    rows = {
        row.source_record_id: row for row in load_source_register_rows(CANONICAL_WORKBOOK)
    }

    nepa_row = rows["FED-001"]
    assert nepa_row.authority_document_id == "authority_document:nepa-act"
    assert nepa_row.authority_document_class_id == "authority_document"
    assert nepa_row.jurisdiction_scope_id == "scope:federal-us"
    assert nepa_row.parser_admission_class == "structured_web_source"
    assert nepa_row.expected_parser == "html"

    handbook_row = rows["SUP-003"]
    assert handbook_row.authority_document_id == "authority_document:fsh-1909-15"

    pdf_row = rows["WILD-ESA-087"]
    assert pdf_row.parser_admission_class == "direct_document"
    assert pdf_row.expected_parser == "pdf"

    doc_row = rows["R1-021"]
    assert doc_row.parser_admission_class == "direct_document"
    assert doc_row.expected_parser == "doc"

    image_row = rows["WILD-ESA-094"]
    assert image_row.parser_admission_class == "direct_document"
    assert image_row.expected_parser == "image"

    eo_row = rows["FED-042"]
    assert eo_row.parser_admission_class == "structured_web_source"
    assert eo_row.expected_parser == "html"

    conformity_row = rows["FED-044"]
    assert conformity_row.parser_admission_class == "structured_web_source"
    assert conformity_row.expected_parser == "html"
    assert conformity_row.authority_document_class_id == "authority_document"
    assert conformity_row.jurisdiction_scope_id == "scope:federal-us"
    assert conformity_row.authority_section_id == (
        "authority_document:federal-united-states-regulation-40-cfr-part-93-subpart-b-general-conformity"
        "#section:40-cfr-part-93-subpart-b"
    )

    cwa_401_row = rows["FED-045"]
    assert cwa_401_row.parser_admission_class == "structured_web_source"
    assert cwa_401_row.expected_parser == "html"
    assert cwa_401_row.authority_document_class_id == "authority_document"
    assert cwa_401_row.jurisdiction_scope_id == "scope:federal-us"

    nhpa_110_row = rows["FED-052"]
    assert nhpa_110_row.parser_admission_class == "structured_web_source"
    assert nhpa_110_row.expected_parser == "html"
    assert nhpa_110_row.authority_document_class_id == "authority_document"
    assert nhpa_110_row.jurisdiction_scope_id == "scope:federal-us"

    arpa_uniform_row = rows["FED-053"]
    assert arpa_uniform_row.parser_admission_class == "structured_web_source"
    assert arpa_uniform_row.expected_parser == "html"
    assert arpa_uniform_row.authority_document_class_id == "authority_document"
    assert arpa_uniform_row.jurisdiction_scope_id == "scope:federal-us"

    nagpra_regulations_row = rows["FED-054"]
    assert nagpra_regulations_row.parser_admission_class == "structured_web_source"
    assert nagpra_regulations_row.expected_parser == "html"
    assert nagpra_regulations_row.authority_document_class_id == "authority_document"
    assert nagpra_regulations_row.jurisdiction_scope_id == "scope:federal-us"

    corps_regulatory_row = rows["FED-049"]
    assert corps_regulatory_row.parser_admission_class == "structured_web_source"
    assert corps_regulatory_row.expected_parser == "html"
    assert corps_regulatory_row.authority_document_class_id == "authority_document"
    assert corps_regulatory_row.jurisdiction_scope_id == "scope:federal-us"

    montana_401_row = rows["STP-031"]
    assert montana_401_row.parser_admission_class == "web_source"
    assert montana_401_row.expected_parser == "html"
    assert montana_401_row.authority_document_class_id == "authority_document"
    assert montana_401_row.jurisdiction_scope_id == "scope:ea-project-review"

    washington_dahp_row = rows["STP-035"]
    assert washington_dahp_row.parser_admission_class == "web_source"
    assert washington_dahp_row.expected_parser == "html"
    assert washington_dahp_row.authority_document_class_id == "authority_document"
    assert washington_dahp_row.jurisdiction_scope_id == "scope:ea-project-review"

    sacred_sites_row = rows["FED-039"]
    assert sacred_sites_row.parser_admission_class == "structured_web_source"
    assert sacred_sites_row.expected_parser == "html"

    us_code_row = rows["FED-029"]
    assert us_code_row.parser_admission_class == "structured_web_source"
    assert us_code_row.expected_parser == "html"

    eagle_permits_row = rows["FED-060"]
    assert eagle_permits_row.parser_admission_class == "structured_web_source"
    assert eagle_permits_row.expected_parser == "html"
    assert eagle_permits_row.authority_document_class_id == "authority_document"
    assert eagle_permits_row.jurisdiction_scope_id == "scope:federal-us"

    efh_statute_row = rows["FED-061"]
    assert efh_statute_row.parser_admission_class == "structured_web_source"
    assert efh_statute_row.expected_parser == "html"
    assert efh_statute_row.authority_document_class_id == "authority_document"
    assert efh_statute_row.jurisdiction_scope_id == "scope:federal-us"

    efh_regulation_row = rows["FED-062"]
    assert efh_regulation_row.parser_admission_class == "structured_web_source"
    assert efh_regulation_row.expected_parser == "html"
    assert efh_regulation_row.authority_document_class_id == "authority_document"
    assert efh_regulation_row.jurisdiction_scope_id == "scope:federal-us"

    ncp_regulation_row = rows["FED-063"]
    assert ncp_regulation_row.parser_admission_class == "structured_web_source"
    assert ncp_regulation_row.expected_parser == "html"
    assert ncp_regulation_row.authority_document_class_id == "authority_document"
    assert ncp_regulation_row.jurisdiction_scope_id == "scope:federal-us"

    farmland_statute_row = rows["FED-064"]
    assert farmland_statute_row.parser_admission_class == "structured_web_source"
    assert farmland_statute_row.expected_parser == "html"
    assert farmland_statute_row.authority_document_class_id == "authority_document"
    assert farmland_statute_row.jurisdiction_scope_id == "scope:federal-us"

    farmland_regulation_row = rows["FED-065"]
    assert farmland_regulation_row.parser_admission_class == "structured_web_source"
    assert farmland_regulation_row.expected_parser == "html"
    assert farmland_regulation_row.authority_document_class_id == "authority_document"
    assert farmland_regulation_row.jurisdiction_scope_id == "scope:federal-us"

    drinking_water_statute_row = rows["FED-066"]
    assert drinking_water_statute_row.parser_admission_class == "structured_web_source"
    assert drinking_water_statute_row.expected_parser == "html"
    assert drinking_water_statute_row.authority_document_class_id == "authority_document"
    assert drinking_water_statute_row.jurisdiction_scope_id == "scope:federal-us"

    uic_regulation_row = rows["FED-067"]
    assert uic_regulation_row.parser_admission_class == "structured_web_source"
    assert uic_regulation_row.expected_parser == "html"
    assert uic_regulation_row.authority_document_class_id == "authority_document"
    assert uic_regulation_row.jurisdiction_scope_id == "scope:federal-us"

    fifra_guidance_row = rows["FED-068"]
    assert fifra_guidance_row.parser_admission_class == "web_source"
    assert fifra_guidance_row.expected_parser == "html"
    assert fifra_guidance_row.authority_document_class_id == "authority_document"
    assert fifra_guidance_row.jurisdiction_scope_id == "scope:federal-us"

    invasive_species_row = rows["FED-069"]
    assert invasive_species_row.parser_admission_class == "structured_web_source"
    assert invasive_species_row.expected_parser == "html"
    assert invasive_species_row.authority_document_class_id == "authority_document"
    assert invasive_species_row.jurisdiction_scope_id == "scope:federal-us"

    minerals_overlay_row = rows["FED-070"]
    assert minerals_overlay_row.parser_admission_class == "structured_web_source"
    assert minerals_overlay_row.expected_parser == "html"
    assert minerals_overlay_row.authority_document_class_id == "authority_document"
    assert minerals_overlay_row.jurisdiction_scope_id == "scope:federal-us"

    hfra_chapter_row = rows["FED-071"]
    assert hfra_chapter_row.parser_admission_class == "structured_web_source"
    assert hfra_chapter_row.expected_parser == "html"
    assert hfra_chapter_row.authority_document_class_id == "authority_document"
    assert hfra_chapter_row.jurisdiction_scope_id == "scope:federal-us"

    hfra_admin_review_row = rows["FED-072"]
    assert hfra_admin_review_row.parser_admission_class == "structured_web_source"
    assert hfra_admin_review_row.expected_parser == "html"
    assert hfra_admin_review_row.authority_document_class_id == "authority_document"
    assert hfra_admin_review_row.jurisdiction_scope_id == "scope:federal-us"

    wildfire_resilience_row = rows["FED-073"]
    assert wildfire_resilience_row.parser_admission_class == "structured_web_source"
    assert wildfire_resilience_row.expected_parser == "html"
    assert wildfire_resilience_row.authority_document_class_id == "authority_document"
    assert wildfire_resilience_row.jurisdiction_scope_id == "scope:federal-us"

    sage_grouse_ce_row = rows["FED-074"]
    assert sage_grouse_ce_row.parser_admission_class == "structured_web_source"
    assert sage_grouse_ce_row.expected_parser == "html"
    assert sage_grouse_ce_row.authority_document_class_id == "authority_document"
    assert sage_grouse_ce_row.jurisdiction_scope_id == "scope:federal-us"

    fuel_break_row = rows["FED-075"]
    assert fuel_break_row.parser_admission_class == "structured_web_source"
    assert fuel_break_row.expected_parser == "html"
    assert fuel_break_row.authority_document_class_id == "authority_document"
    assert fuel_break_row.jurisdiction_scope_id == "scope:federal-us"

    emergency_actions_row = rows["FED-076"]
    assert emergency_actions_row.parser_admission_class == "structured_web_source"
    assert emergency_actions_row.expected_parser == "html"
    assert emergency_actions_row.authority_document_class_id == "authority_document"
    assert emergency_actions_row.jurisdiction_scope_id == "scope:federal-us"

    timber_overlay_row = rows["FED-077"]
    assert timber_overlay_row.parser_admission_class == "structured_web_source"
    assert timber_overlay_row.expected_parser == "html"
    assert timber_overlay_row.authority_document_class_id == "authority_document"
    assert timber_overlay_row.jurisdiction_scope_id == "scope:federal-us"

    seven_county_row = rows["FED-078"]
    assert seven_county_row.parser_admission_class == "direct_document"
    assert seven_county_row.expected_parser == "pdf"
    assert seven_county_row.authority_document_class_id == "authority_document"
    assert seven_county_row.jurisdiction_scope_id == "scope:federal-us"

    apa_reviewability_row = rows["FED-079"]
    assert apa_reviewability_row.parser_admission_class == "structured_web_source"
    assert apa_reviewability_row.expected_parser == "html"
    assert apa_reviewability_row.authority_document_class_id == "authority_document"
    assert apa_reviewability_row.jurisdiction_scope_id == "scope:federal-us"

    directives_notice_row = rows["FED-080"]
    assert directives_notice_row.parser_admission_class == "direct_document"
    assert directives_notice_row.expected_parser == "pdf"
    assert directives_notice_row.authority_document_class_id == "authority_document"
    assert directives_notice_row.jurisdiction_scope_id == "scope:federal-us"

    musya_row = rows["FED-081"]
    assert musya_row.parser_admission_class == "structured_web_source"
    assert musya_row.expected_parser == "html"
    assert musya_row.authority_document_class_id == "authority_document"
    assert musya_row.jurisdiction_scope_id == "scope:federal-us"

    organic_purposes_row = rows["FED-082"]
    assert organic_purposes_row.parser_admission_class == "structured_web_source"
    assert organic_purposes_row.expected_parser == "html"
    assert organic_purposes_row.authority_document_class_id == "authority_document"
    assert organic_purposes_row.jurisdiction_scope_id == "scope:federal-us"

    cave_resources_row = rows["FED-083"]
    assert cave_resources_row.parser_admission_class == "direct_document"
    assert cave_resources_row.expected_parser == "pdf"
    assert cave_resources_row.authority_document_class_id == "authority_document"
    assert cave_resources_row.jurisdiction_scope_id == "scope:federal-us"

    paleontology_row = rows["FED-084"]
    assert paleontology_row.parser_admission_class == "direct_document"
    assert paleontology_row.expected_parser == "pdf"
    assert paleontology_row.authority_document_class_id == "authority_document"
    assert paleontology_row.jurisdiction_scope_id == "scope:federal-us"

    wsr_regulation_row = rows["FED-085"]
    assert wsr_regulation_row.parser_admission_class == "direct_document"
    assert wsr_regulation_row.expected_parser == "pdf"
    assert wsr_regulation_row.authority_document_class_id == "authority_document"
    assert wsr_regulation_row.jurisdiction_scope_id == "scope:federal-us"

    trails_row = rows["FED-086"]
    assert trails_row.parser_admission_class == "structured_web_source"
    assert trails_row.expected_parser == "html"
    assert trails_row.authority_document_class_id == "authority_document"
    assert trails_row.jurisdiction_scope_id == "scope:federal-us"

    montana_wsa_row = rows["FED-087"]
    assert montana_wsa_row.parser_admission_class == "direct_document"
    assert montana_wsa_row.expected_parser == "pdf"
    assert montana_wsa_row.authority_document_class_id == "authority_document"
    assert montana_wsa_row.jurisdiction_scope_id == "scope:federal-us"

    region_forest_index_row = rows["R1PLAN-region-1-northern-region-02"]
    assert region_forest_index_row.parser_admission_class == "structured_web_source"
    assert region_forest_index_row.expected_parser == "html"
    assert region_forest_index_row.authority_document_class_id == "authority_document"
    assert region_forest_index_row.jurisdiction_scope_id == "scope:usfs-region-1"

    beaverhead_planning_row = rows["R1PLAN-beaverhead-deerlodge-nf-01"]
    assert beaverhead_planning_row.parser_admission_class == "structured_web_source"
    assert beaverhead_planning_row.expected_parser == "html"
    assert beaverhead_planning_row.authority_document_class_id == "authority_document"
    assert beaverhead_planning_row.jurisdiction_scope_id == "scope:region1-forest-unit"

    challenge_repair_row = rows["FPS-344"]
    assert challenge_repair_row.parser_admission_class == "structured_web_source"
    assert challenge_repair_row.expected_parser == "html"

    directives_row = rows["USDA-008"]
    assert directives_row.parser_admission_class == "structured_web_source"
    assert directives_row.expected_parser == "html"

    accessibility_row = rows["USDA-009"]
    assert accessibility_row.parser_admission_class == "structured_web_source"
    assert accessibility_row.expected_parser == "html"

    foia_row = rows["USDA-010"]
    assert foia_row.parser_admission_class == "structured_web_source"
    assert foia_row.expected_parser == "html"

    tribal_row = rows["USDA-011"]
    assert tribal_row.parser_admission_class == "direct_document"
    assert tribal_row.expected_parser == "pdf"

    nwcg_row = rows["PROG-010"]
    assert nwcg_row.parser_admission_class == "direct_document"
    assert nwcg_row.expected_parser == "pdf"

    scc_workbook_row = rows["R1-SCC-NPC-004"]
    assert scc_workbook_row.parser_admission_class == "direct_document"
    assert scc_workbook_row.expected_parser == "xlsx"

    nondiscrimination_row = rows["USDA-012"]
    assert nondiscrimination_row.parser_admission_class == "structured_web_source"
    assert nondiscrimination_row.expected_parser == "html"

    info_quality_row = rows["USDA-013"]
    assert info_quality_row.parser_admission_class == "structured_web_source"
    assert info_quality_row.expected_parser == "html"

    forest_plan_row = rows["FPS-002"]
    assert forest_plan_row.authority_document_class_id == "forest_plan"
    assert forest_plan_row.jurisdiction_scope_id == "scope:region1-forest-unit"
    assert forest_plan_row.authority_section_id is not None


def test_source_register_loader_projects_governed_lineage_metadata_into_workbook_sources() -> None:
    config = load_config(CONFIG)
    sources = {
        source.source_record_id: source
        for source in load_canonical_sources(CANONICAL_WORKBOOK, config.workbook)
    }

    usfs_026 = sources["USFS-026"]

    assert usfs_026.metadata["supersession_status"] == "superseded by current authority"
    assert usfs_026.metadata["supersession_status_id"] == "superseded_by_current_authority"
    assert usfs_026.metadata["replacement_source_record_ids"] == "USFS-023"
    assert "removed from the directive system" in usfs_026.metadata["source_currentness_status"]


def test_source_register_loader_rejects_blocked_alias_without_context() -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        workbook_path = Path(tmp_dir) / "ambiguous-source-register.xlsx"
        workbook = load_workbook(CANONICAL_WORKBOOK)
        sheet = workbook["Document_Register_Master"]
        headers = {
            str(sheet.cell(4, column).value).strip(): column
            for column in range(1, sheet.max_column + 1)
        }
        target_row = 5
        sheet.cell(target_row, headers["Document_Title"]).value = "Forest Plan"
        sheet.cell(target_row, headers["Citation_or_Code"]).value = ""
        sheet.cell(target_row, headers["Issuing_Entity"]).value = ""
        sheet.cell(target_row, headers["Jurisdiction_or_Unit"]).value = ""
        sheet.cell(target_row, headers["Issue_or_Effective_Date"]).value = ""
        workbook.save(workbook_path)

        try:
            load_source_register_rows(workbook_path)
        except ValueError as exc:
            assert "requires more context before resolving blocked alias" in str(exc)
        else:
            raise AssertionError("Expected canonical loader to fail on blocked alias without context")
